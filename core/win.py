import os
import re
from typing import Literal
from shutil import rmtree
from pathlib import Path

import dill
import seismicutils as su
import numpy as np
import pyqtgraph as pg
import openpyxl as px
from openpyxl.worksheet.worksheet import Worksheet
from PyQt5.QtGui import QIntValidator, QDoubleValidator, QFont, QColor
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QPoint
from PyQt5.QtWidgets import QApplication, QMessageBox, QFileDialog, QDialog, QTableWidgetItem, QMainWindow, QMenu, QTableWidgetItem, QHeaderView, QTableWidget, QLabel

import core
from ui.main_win import Ui_MainWindow
from ui.win_importGM import Ui_win_importGM
from ui.win_importGM1 import Ui_win_importGM1
from ui.win_mass import Ui_win_mass
from ui.win_mat import Ui_win_mat
from ui.win_solve_setting import Ui_win_solve_setting
from ui.win_OSmat import Ui_win_OSmat
from ui.win_run import Ui_win_run
from ui.win_tcl_file import Ui_win_tcl_file
from ui.win_data import Ui_win_data
from ui.win_export import Ui_win_export
from ui.win_about import Ui_win_about


SOFTWARE = '非线性多自由度时程分析软件'
VERSION = 'V2.0'
DATE = '2024.08.24'
TEMP_PATH = Path(os.getenv('TEMP')).as_posix()
ROOT = Path.cwd()


class MyWin(QMainWindow):
    g = 9800
    unit_SF = [1, 1 / g, 10 / g, 1000 / g]
    unit = ['g', 'mm/s^2', 'cm/s^2', 'm/s^2']
    setting1 = ['Plain', 'Lagrange', 'Penalty', 'Transformation']
    setting2 = ['Plain', 'RCM', 'AMD']
    setting3 = ['BandGeneral', 'BandSPD', 'ProfileSPD', 'SuperLU', 'UmfPack', 'FullGeneral', 'SparseSYM']
    setting4 = ['NormUnbalance', 'NormDispIncr', 'EnergyIncr', 'RelativeNormUnbalance',
                'RelativeNormDispIncr', 'RelativeTotalNormDispIncr', 'RelativeEnergyIncr', 'FixedNumIter']
    setting5 = ['Linear', 'Newton', 'NewtonLineSearch', 'ModifiedNewton', 'KrylovNewton',
                'SecantNewton', 'BFGS', 'Broyden']
    setting6 = ['CentralDifference', 'Newmark', 'HHT', 'GeneralizedAlpha', 'TRBDF2', 'Explicitdifference']
    print_result = False

    def __init__(self, test: bool=False):
        super().__init__()
        self.init_pg()
        self.init_ui()
        self.init_gm_var()
        self.init_var()
        self.init_result()
        if test:
            self.init_test()

    def init_pg(self):
        """初始化pyqtgraph画笔"""
        self.pen1 = pg.mkPen(color=(68, 114, 196), width=2)  # matplotlib蓝色:30, 120, 180
        self.pen2 = pg.mkPen(color='orange', width=2)  # 橙色
        self.pen3 = pg.mkPen(color='green', width=2)  # 绿色
        self.pen4 = pg.mkPen(color=(211, 211, 211), width=2)  # 亮灰色
        self.pen5 = pg.mkPen(color='red', width=2)  # 红色
        self.pen_axis = pg.mkPen(color='black', width=1)  # 坐标轴
        self.pen_grid = pg.mkPen(color='grey', width=1)
        self.font_axis = QFont('Arial')
        self.font_axis.setPixelSize(16)  # 坐标轴字体大小

    def init_ui(self):
        """初始化ui"""
        # tab 1
        self.ui = Ui_MainWindow()  # 主窗口
        self.ui.setupUi(self)
        self.setWindowTitle(f'{SOFTWARE} {VERSION}')
        self.ui.pushButton.clicked.connect(self.open_win_select_gm)
        self.ui.pushButton_2.clicked.connect(self.open_win_select_gm1)
        self.ui.pushButton_5.clicked.connect(self.delete_seleted)
        self.ui.pushButton_4.clicked.connect(self.delete_all)
        self.ui.listWidget.itemClicked.connect(self.plot_gm)
        self.pg1 = self.replace_to_pyqtgraph(self.ui.graphicsView, self.ui.verticalLayout_4, 2)
        self.pg1.setLabel(axis='bottom', text='t [s]')
        self.pg1.setLabel(axis='left', text='a [g]')
        self.pg1.getAxis('bottom').label.setFont(pg.QtGui.QFont(self.font_axis))
        self.pg1.getAxis('left').label.setFont(pg.QtGui.QFont(self.font_axis))
        validator = QDoubleValidator(0, 10000, 7)
        validator.setNotation(QDoubleValidator.StandardNotation)
        self.ui.lineEdit_4.setValidator(validator)
        self.ui.lineEdit.setValidator(validator)
        self.ui.pushButton_3.clicked.connect(self.update_gm)
        # tab 2
        validator1 = QIntValidator()
        self.ui.lineEdit_2.setValidator(validator1)
        self.ui.lineEdit_2.setPlaceholderText('1~10000')
        self.ui.lineEdit_2.editingFinished.connect(self.set_N)
        self.ui.pushButton_14.clicked.connect(self.open_win_mass)
        self.ui.pushButton_9.clicked.connect(lambda: self.open_win_define_mat(None, 0))
        self.ui.pushButton_7.clicked.connect(lambda: self.open_win_define_mat(None, 1))
        self.ui.pushButton_8.clicked.connect(lambda: self.open_win_define_mat(None, 2))
        self.ui.pushButton_6.clicked.connect(lambda: self.open_win_define_mat(None, 3))
        self.ui.pushButton_10.clicked.connect(lambda: self.open_win_define_mat(None, 4))
        self.ui.listWidget_2.itemDoubleClicked.connect(self.open_win_define_mat)
        self.ui.pushButton_12.clicked.connect(self.delete_seleted_mat)
        self.ui.pushButton_13.clicked.connect(self.delete_all_mat)
        self.ui.listWidget_3.itemClicked.connect(self.clicked_story)
        self.ui.pushButton_17.clicked.connect(self.assign_mat_to_story)
        self.ui.pushButton_15.clicked.connect(self.delete_story_mat)
        self.ui.pushButton_16.clicked.connect(self.delete_all_story_mat)
        self.ui.lineEdit_2.setPlaceholderText('1~10000')
        self.ui.lineEdit_6.setValidator(validator)
        self.ui.lineEdit_6.editingFinished.connect(self.set_fvtime)
        self.ui.pushButton_11.clicked.connect(self.define_OSmat)
        # tab 3
        self.ui.tabWidget.currentChanged.connect(self.init_tab3)
        self.ui.radioButton_2.toggled.connect(lambda: self.ui.pushButton_18.setEnabled(self.ui.radioButton_2.isChecked()))
        self.ui.radioButton_3.toggled.connect(lambda: self.ui.comboBox_3.setEnabled(not self.ui.radioButton_3.isChecked()))
        self.ui.radioButton_3.toggled.connect(lambda: self.ui.comboBox_4.setEnabled(not self.ui.radioButton_3.isChecked()))
        self.ui.radioButton_3.toggled.connect(lambda: self.ui.lineEdit_3.setEnabled(not self.ui.radioButton_3.isChecked()))
        self.ui.radioButton_3.toggled.connect(lambda: self.ui.lineEdit_5.setEnabled(not self.ui.radioButton_3.isChecked()))
        self.ui.radioButton_3.toggled.connect(self.clicked_Rayleigh)
        self.ui.lineEdit_3.setValidator(validator)
        self.ui.lineEdit_5.setValidator(validator)
        self.ui.lineEdit_3.editingFinished.connect(lambda: self.zeta_edit_finished(self.ui.lineEdit_3))
        self.ui.lineEdit_5.editingFinished.connect(lambda: self.zeta_edit_finished(self.ui.lineEdit_5))
        self.ui.pushButton_18.clicked.connect(self.choose_OS_terminal)
        self.ui.pushButton_20.clicked.connect(self.clicked_build_tcl_file)
        self.ui.pushButton_19.clicked.connect(self.run)
        self.pg3 = self.replace_to_pyqtgraph(self.ui.graphicsView_3, self.ui.verticalLayout_12, 1)
        self.ui.comboBox_6.activated.connect(self.update_hyeteretic_curve_list)
        self.ui.comboBox_5.activated.connect(self.update_result_combobox)
        self.ui.comboBox_5.activated.connect(self.plot_results)
        self.ui.comboBox_6.activated.connect(self.plot_results)
        self.ui.comboBox_7.activated.connect(self.plot_results)
        self.ui.comboBox_8.activated.connect(self.plot_results)
        self.ui.pushButton_21.clicked.connect(self.show_graph_data)
        self.ui.pushButton_22.clicked.connect(self.export_data)
        # window
        self.ui.action.triggered.connect(self.setting_clicked)
        self.statusBar_label_left = QLabel(f'{SOFTWARE} {VERSION}', self)
        self.statusBar().addWidget(self.statusBar_label_left)
        self.statusBar_label_right = QLabel('', self)
        self.statusBar().addPermanentWidget(self.statusBar_label_right)
        self.ui.action_2.triggered.connect(self.open_win_about)


    def init_gm_var(self):
        """初始化地震动变量"""
        self.gm = []  # 加速度序列
        self.gm_name = []
        self.gm_N = 0
        self.gm_dt  = []
        self.gm_NPTS = []
        self.gm_t = []  # 时间序列
        self.gm_duration = []  # 持时
        self.gm_unit = []  # 单位
        self.gm_PGA = []

    def init_var(self):
        self.TEMP_PATH = os.getenv('TEMP')  # 临时文件路径
        self.current_gm_idx = None  # 当前选择地震动的序号
        self.N = 0  # 自由度数量
        self.m = []  # 各自由度质量
        self.mat_lib = []  # 备选材料 [['备注名', 0(材料种类(常用材料:0-5,OS材料为-1)), 0(是否有备注名), 'Steel01', 1, 235, 206000, 0.02], ...]
        self.mat_N = 0  # 当前材料数量
        self.story_mat = []  # 各层材料 [[1, 2](材料编号)], ...]
        self.current_story = 0  # 当前选中层数
        self.mode_num = 0  # 最大有效振型数
        self.fvtime = 0  # 自由振动时长
        self.has_damping = True  # 是否有阻尼
        self.setting = [3, 0, 0, 0, 1, 1, '', '', '1e-5', '60', '0.5', '0.25', '1', '1e-6', '1']
        self.setting_default = [3, 0, 0, 0, 1, 1, '', '', '1e-5', '60', '0.5', '0.25', '1', '1e-6', '1']
        self.OS_terminal = None  # OpenSees求解器路径
        self.current_plot_data = None  # 当前绘制的图像的数据
        self.export_type = None  # 导出数据的类型
        self.has_os_terminal = False  # 是否有OpenSees求解器路径
        
    def init_result(self):
        """初始化计算结果"""
        self.result_exists = False
        self.result_T = None
        self.result_mode = None

    def replace_to_pyqtgraph(self, graphicsView, layout, index):
        """将graphicsView控件替换为pyqtgrapg"""
        layout.removeWidget(graphicsView)
        graphicsView.deleteLater()
        pg_widget = pg.PlotWidget()
        pg_widget.setBackground('w')
        # 显示上轴和右轴
        pg_widget.showAxis('top')
        pg_widget.showAxis('right')
        # 设置坐标轴宽度
        pg_widget.getAxis('bottom').setPen(self.pen_axis)
        pg_widget.getAxis('top').setPen(self.pen_axis)
        pg_widget.getAxis('left').setPen(self.pen_axis)
        pg_widget.getAxis('right').setPen(self.pen_axis)
        # 更改坐标轴刻度字体大小
        pg_widget.getAxis('bottom').setTickFont(self.font_axis)
        pg_widget.getAxis('top').setTickFont(self.font_axis)
        pg_widget.getAxis('left').setTickFont(self.font_axis)
        pg_widget.getAxis('right').setTickFont(self.font_axis)
        # 更改坐标轴字体颜色
        pg_widget.getAxis('bottom').setTextPen(QColor('black'))
        pg_widget.getAxis('top').setTextPen(QColor('black'))
        pg_widget.getAxis('left').setTextPen(QColor('black'))
        pg_widget.getAxis('right').setTextPen(QColor('black'))
        # 不显示上轴和右轴的刻度标签
        pg_widget.getAxis('top').setTicks([])
        pg_widget.getAxis('right').setTicks([])
        # 显示网格线
        pg_widget.showGrid(True, True)
        # 添加到原layout中
        layout.insertWidget(index, pg_widget)
        return pg_widget

    def init_test(self):
        """软件调试模式，将自动导入自动波和设置模型"""
        self.N = 3
        th1 = np.loadtxt('data/ChiChi.dat')[:, 1]
        th2 = np.loadtxt('data/Kobe.dat')[:, 1]
        self.gm.append(th1)
        self.gm.append(th2)
        self.gm_name.append('ChiChi')
        self.gm_name.append('Kobe')
        self.gm_N = 2
        self.gm_dt.append(0.01)
        self.gm_dt.append(0.01)
        self.gm_NPTS.append(5279)
        self.gm_NPTS.append(4091)
        self.gm_t.append(np.arange(0, 52.79, 0.01))
        self.gm_t.append(np.arange(0, 40.91, 0.01))
        self.gm_duration.append(52.79)
        self.gm_duration.append(40.91)
        self.gm_unit.append('g')
        self.gm_unit.append('g')
        self.gm_PGA.append(max(abs(th1)))
        self.gm_PGA.append(max(abs(th2)))
        self.N = 3
        self.m = [2, 1, 1]
        self.mat_lib.append(['Steel01-1', -1, 1, 'Steel01', 1, 8000, 8000, 0.02])
        self.mat_lib.append(['Steel01-2', -1, 1, 'Steel01', 2, 5000, 5000, 0.02])
        self.story_mat = [[1, 2], [1], [1]]
        self.mode_num = 3
        self.OS_terminal = r'D:\OpenSees3.5.0\bin\OpenSees.exe'
        self.ui.lineEdit_2.setText('3')


# ----------------------------- Static methods --------------------------------

    @staticmethod
    def add_free_vibration(th: np.ndarray, dt: float, fv_time: float):
        if fv_time == 0:
            return th
        n = int(fv_time / dt)  # 需要补0的个数
        th_0 = np.zeros(n)
        th = np.append(th, th_0)
        return th


# ----------------------------- tab 1 ------------------------------------------

    def open_win_select_gm(self):
        win = Win_importGM(self)
        win.exec_()

    def gm_list_update(self):
        """更新地震动列表"""
        self.ui.listWidget.clear()
        # 1 更新地震动列表
        for i, name in enumerate(self.gm_name):
            self.ui.listWidget.addItem(f'({i+1}) {name}')

    def delete_seleted(self):
        """删除选中地震动"""
        total_num = self.ui.listWidget.count()
        if total_num == 0 or len(self.ui.listWidget.selectedIndexes()) == 0:
            return 0
        idx = self.ui.listWidget.selectedIndexes()[0].row()
        del self.gm[idx]
        del self.gm_name[idx]
        self.gm_N -= 1
        del self.gm_dt[idx]
        del self.gm_NPTS[idx]
        del self.gm_t[idx]
        del self.gm_duration[idx]
        self.gm_list_update()
        if idx == self.current_gm_idx:
            self.plot_nothing()
        elif idx < self.current_gm_idx:
            self.current_gm_idx -= 1
            self.plot_gm(None, self.current_gm_idx)

    def delete_all(self):
        """删除所有地震动"""
        total_num = self.ui.listWidget.count()
        if total_num == 0:
            self.plot_nothing()
            return 0
        if QMessageBox.question(self, '提示', '是否删除所有地震动？') == QMessageBox.Yes:
            self.init_gm_var()
            self.gm_list_update()
            self.plot_nothing()

    def plot_gm(self, item, idx=None):
        """展示地震动信息并绘图（双击地震动后触发）"""
        if type(idx) is not int:
            idx = self.ui.listWidget.row(item)
            self.ui.label_5.clear()
        self.current_gm_idx = idx
        print('【MyWin, plot_gm】当前地震动idx：', self.current_gm_idx)
        print('【MyWin, plot_gm】len(th) =', len(self.gm[idx]))
        print('【MyWin, plot_gm】总地震动数量 =', self.gm_N)
        print('【MyWin, plot_gm】dt =', self.gm_dt[idx])
        print('【MyWin, plot_gm】NPTS =', self.gm_NPTS[idx])
        print('【MyWin, plot_gm】len(t) =', len(self.gm_t[idx]))
        print('【MyWin, plot_gm】duration =', self.gm_duration[idx])
        print('【MyWin, plot_gm】unit =', self.gm_unit[idx])
        print('【MyWin, plot_gm】PGA =', self.gm_PGA[idx])
        print('------------------------------')
        self.ui.label_12.setText(str(f'({idx+1}) ' + self.gm_name[idx]))
        # self.ui.lineEdit_3.setText(str(self.gm_SF[idx]))
        self.ui.lineEdit_4.setText(str(round(max(abs(self.gm[idx])), 3)))
        self.ui.lineEdit.setText(str(round(self.gm_dt[idx], 5)))
        self.ui.label_8.setText(str(round(self.gm_duration[idx], 5)))
        self.ui.label_9.setText(str(int(self.gm_NPTS[idx] + 1)))
        self.pg1.clear()
        self.pg1.plot(self.gm_t[idx], self.gm[idx], pen=self.pen1, name='加速度时程曲线')
        self.pg1.setLabel(axis='left', text=f'a [{self.ui.comboBox.currentText()}]')
        self.pg1.autoRange()

    def plot_nothing(self):
        self.pg1.clear()
        self.ui.comboBox.setCurrentIndex(0)
        self.ui.label_12.clear()
        self.ui.lineEdit_4.clear()
        self.ui.lineEdit.clear()
        self.ui.label_8.clear()
        self.ui.label_9.clear()
        self.ui.label_5.clear()

    def update_gm(self):
        """更新地震动参数"""
        if type(self.current_gm_idx) is not int:
            return
        if float(self.ui.lineEdit_4.text()) == 0:
            self.ui.lineEdit_4.setText('0.00001')
        if float(self.ui.lineEdit.text()) == 0:
            self.ui.lineEdit.setText('0.00001')
        idx = self.current_gm_idx
        self.gm_PGA[idx] = float(self.ui.lineEdit_4.text())
        self.gm_dt[idx] = float(self.ui.lineEdit.text())
        self.gm_unit[idx] = self.ui.comboBox.currentText()
        self.gm_duration[idx] = self.gm_NPTS[idx] * float(self.ui.lineEdit.text())
        th_old = self.gm[idx]
        th = th_old / max(abs(th_old)) * float(self.ui.lineEdit_4.text())
        t = np.linspace(0, self.gm_NPTS[idx] * float(self.ui.lineEdit.text()), self.gm_NPTS[idx] + 1)
        self.gm[idx] = th
        self.gm_t[idx] = t
        self.ui.label_5.setText(f'{self.gm_name[idx]}已更新')
        self.plot_gm(None, idx=idx)

    def open_win_select_gm1(self):
        win = Win_importGM1(self)
        win.exec_()


# ----------------------------- tab 2 ------------------------------------------

    def set_N(self):
        N = int(self.ui.lineEdit_2.text())
        if N < 1:
            N = 1
        elif N > 10000:
            N = 10000
        self.ui.lineEdit_2.setText(str(N))
        count = len(self.story_mat)  # 楼层材料指派列表的数量
        if N < count and N != self.N:  # 自由度减少
            self.story_mat = self.story_mat[: N]
        elif N > count and N != self.N:  # 自由度增加
            for i in range(count, N):
                self.story_mat.append([])
        self.N = N
        if self.N == 1:
            self.ui.comboBox_4.setEnabled(False)
            self.ui.lineEdit_5.setEnabled(False)
        else:
            if self.ui.radioButton_4.isChecked():
                self.ui.comboBox_4.setEnabled(True)
                self.ui.lineEdit_5.setEnabled(True)
            else:
                self.ui.comboBox_4.setEnabled(False)
                self.ui.lineEdit_5.setEnabled(False) 
        self.update_story_mat_list()

    def open_win_mass(self):
        win = Win_mass(self)
        win.exec_()

    def open_win_define_mat(self, item, i=None):
        if item:
            mod_idx = self.ui.listWidget_2.row(item)
            i = self.mat_lib[mod_idx][1]  # 材料种类
        else:
            mod_idx = None
        if i >= 0:
            win = Win_mat(self, i, mod_idx)
        elif i == -1:
            win = Win_OSmat(self, mod_idx)
        win.exec_()

    def delete_seleted_mat(self):
        total_num = self.ui.listWidget_2.count()
        if total_num == 0 or len(self.ui.listWidget_2.selectedIndexes()) == 0:
            return 0
        idx = self.ui.listWidget_2.selectedIndexes()[0].row()
        print('【MyWin, delete_seleted_mat】删除：', self.mat_lib[idx])
        del self.mat_lib[idx]
        self.mat_N -= 1
        for i, mat in enumerate(self.mat_lib):
            mat[4] = i + 1
        for i, mat in enumerate(self.story_mat):
            for j, tag in enumerate(mat):
                if tag - 1 > idx:
                    self.story_mat[i][j] -= 1
                elif tag - 1 == idx:
                    del self.story_mat[i][j]
        self.update_mat_list()
        self.update_conbeBox_mat()
        self.update_story_mat_list()

    def delete_all_mat(self):
        if self.mat_N == 0:
            return
        if QMessageBox.question(self, '警告', '是否删除所有备选材料？') == QMessageBox.Yes:
            self.mat_lib = []
            self.mat_N = 0
            self.story_mat = [[] for i in range(self.N)]
            self.update_mat_list()
            self.update_conbeBox_mat()
            self.update_story_mat_list()

    def update_mat_list(self):
        """更新已定义材料"""
        if self.mat_N == 0:
            self.ui.listWidget_2.clear()
        else:
            self.ui.listWidget_2.clear()
            for i, mat in enumerate(self.mat_lib):
                self.ui.listWidget_2.addItem(f'({i+1}) {mat[0]}')

    def update_conbeBox_mat(self):
        """更新选择材料以添加的框"""
        if self.mat_N == 0:
            self.ui.comboBox_2.clear()
        else:
            self.ui.comboBox_2.clear()
            for i, mat in enumerate(self.mat_lib):
                self.ui.comboBox_2.addItem(f'({i+1}) {mat[0]}')

    def update_story_mat_list(self, clear_all=False):
        """更新选中该层材料进行修改"""
        self.ui.listWidget_3.clear()
        if clear_all:
            return
        for i, mat in enumerate(self.story_mat):
            if not mat:
                self.ui.listWidget_3.addItem(f'第({i+1})层 未指派材料')
            else:
                self.ui.listWidget_3.addItem(f'第({i+1})层')

    def clicked_story(self, item):
        idx = self.ui.listWidget_3.row(item)
        print(f'【MyWin, clicked_story】已选择第{idx+1}层 (idx={idx})')
        self.ui.textBrowser.setText(f'第{idx+1}层：')
        if self.story_mat[idx]:
            text = f'第({idx+1})层：\n'
            for mat_tag in self.story_mat[idx]:
                text += f'{self.mat_lib[mat_tag - 1][0]}\n'
            self.ui.textBrowser.setText(text)
        self.current_story = idx + 1

    def assign_mat_to_story(self):
        if not self.current_story:
            return
        if self.ui.comboBox_2.count() == 0:
            return
        mat_idx = self.ui.comboBox_2.currentIndex()
        self.story_mat[self.current_story - 1].append(mat_idx + 1)
        item = self.ui.listWidget_3.item(self.current_story - 1)
        self.clicked_story(item)
        self.update_story_mat_list()

    def delete_story_mat(self):
        if not self.current_story:
            return
        self.story_mat[self.current_story - 1] = []
        self.update_story_mat_list()
        item = self.ui.listWidget_3.item(self.current_story - 1)
        self.clicked_story(item)

    def delete_all_story_mat(self):
        if not self.current_story:
            return
        self.story_mat = [[] for _ in range(self.N)]
        self.update_story_mat_list()
        self.ui.textBrowser.clear()

    def set_fvtime(self):
        fv = self.ui.lineEdit_6.text()
        if not fv:
            self.ui.lineEdit_6.setText('0')
            self.fvtime = 0
            return
        self.fvtime = float(fv)

    def define_OSmat(self, item=None):
        win = Win_OSmat(self)
        win.exec_()

# ----------------------------- tab 3 ------------------------------------------

    def init_tab3(self, index):
        if index != 2:
            return
        print('【MyWin, init_tab3】初始化tab3')
        if not self.model_is_complete():
            return
        mode_num = 0
        for mi in self.m:
            if mi > 0:
                mode_num += 1
        self.mode_num = min(mode_num, 5)  # 最大有效振型数
        self.ui.comboBox_3.clear()
        self.ui.comboBox_4.clear()
        for i in range(self.mode_num):
            self.ui.comboBox_3.addItem(f'{i+1}阶')
            self.ui.comboBox_4.addItem(f'{i+1}阶')
        if self.ui.comboBox_4.count() > 1:
            self.ui.comboBox_4.setCurrentIndex(1)

    def clicked_Rayleigh(self):
        if self.N == 1:
            self.ui.comboBox_4.setEnabled(False)
            self.ui.lineEdit_5.setEnabled(False)
        if self.ui.radioButton_4.isChecked():
            self.has_damping = True
        else:
            self.has_damping = False

    def zeta_edit_finished(self, widget):
        if float(widget.text()) > 1:
            widget.setText('1')

    def model_is_complete(self):
        # 判断模型是否完备
        if not all([self.N, self.m]):
            print('【MyWin, model_is_complete】模型不完备，缺少层数或质量')
            return False  # 必须同时定义了层数、质量和地震动
        if self.N != len(self.m):
            print('【MyWin, model_is_complete】质量数量与层数不等')
            return False
        print('【MyWin, model_is_complete】模型完备！')
        print('【MyWin, model_is_complete】', self.mat_lib)
        print('【MyWin, model_is_complete】', self.story_mat)
        return True
        
    def mat_is_complete(self):
        """判断各层是否均赋予材料"""
        if not self.model_is_complete():
            return False
        if len(self.story_mat) != self.N:
            raise ValueError('【MyWin, mat_is_complete】story_mat数量与层数N不等！')
        for i in range(self.N):
            if not self.story_mat[i]:
                print(f'【MyWin, mat_is_complete】第{i+1}层未指派材料')
                QMessageBox.critical(self, '错误', f'第{i+1}层未指派材料！')
                return False
        return True
    
    def gm_is_complete(self):
        """判断地震动是否导入"""
        if not self.gm_N:
            print('【MyWin, gm_is_complete】地震动未导入')
            QMessageBox.critical(self, '错误', '地震动未导入！')
            return False
        return True
    
    def damping_is_correct(self):
        """判断阻尼是否有效"""
        if self.ui.radioButton_3.isChecked():
            return True
        if self.ui.comboBox_3.currentIndex() == self.ui.comboBox_4.currentIndex() and self.N > 1:
            QMessageBox.critical(self, '错误', 'Rayleigh阻尼所选振型不能一致！')
            return False
        else:
            return True
    
    def ready_to_run(self):
        """判断模型是否可以开始求解"""
        if not all([self.model_is_complete(), self.mat_is_complete(),
                    self.gm_is_complete(), self.damping_is_correct()]):
            return False
        if not self.ui.lineEdit_2.text():
            QMessageBox.warning(self, '错误', '自由度数量未输入！')
            return False
        if not self.ui.lineEdit_6.text():
            QMessageBox.warning(self, '错误', '自由振动时长未输入！')
            return False
        if self.ui.radioButton_2.isChecked() and not self.has_os_terminal:
            QMessageBox.warning(self, '错误', '未选择有效的OpenSees.exe！')
            return False
        print('【MyWin, ready_to_run】模型已完备')
        return True

    def run(self, script_type: Literal['py', 'tcl']):
        """script_type: 'py' or 'tcl'"""
        if self.ready_to_run():
            if not (Path(TEMP_PATH) / 'temp_NLMDOF_results').exists():
                os.makedirs((Path(TEMP_PATH) / 'temp_NLMDOF_results').as_posix())
            if self.ui.radioButton.isChecked():
                script_type = 'py'
            else:
                script_type = 'tcl'
            if script_type == 'tcl' and self.OS_terminal is None:
                QMessageBox.warning(self, '警告', '请选择OpenSees.exe')
                return
            self.zeta_mode = [self.ui.comboBox_3.currentIndex() + 1, self.ui.comboBox_4.currentIndex() + 1]
            self.zeta = [self.ui.lineEdit_3.text(), self.ui.lineEdit_3.text()]
            win = Win_run(self, script_type)
            win.signal_converge_fail.connect(self.converge_fail)
            win.signal_finished.connect(self.running_finished)
            win.exec_()
        # else:
        #     QMessageBox.warning(self, '警告', '模型未完全定义！')

    def converge_fail(self):
        pass

    def choose_OS_terminal(self):
        self.OS_terminal = QFileDialog.getOpenFileName(self, '选择OpenSees.exe', '.', 'OpenSees.exe (*.exe)')[0]
        self.OS_terminal = self.OS_terminal.replace('\\', '/')
        print(f'【MyWin, choose_OS_terminal】OpenSees.exe: {self.OS_terminal}')
        if not self.OS_terminal:
            print('【MyWin, choose_OS_terminal】没有选择OpenSees求解器。')
            return
        path1 = '/'.join(self.OS_terminal.split('/')[:-1])  # 父级目录
        test_tcl_text = fr"""set path "{path1}"
set result [catch {{
    wipe
    model basic -ndm 2 -ndf 3
}} error_message]
if {{$result != 0}} {{
    puts "Error"
    puts [open "$path\\ok.txt" w] "fail"
}} else {{
    puts "Ok"
    puts [open "$path\\ok.txt" w] "ok"
}}"""
        try:
            with open(f"{path1}/OS_terminal_test.tcl", 'w') as f:
                f.write(test_tcl_text)
        except:
            path_temp = f"{path1}/OS_terminal_test.tcl"
            print(f'【MyWin, choose_OS_terminal】写入"{path_temp}"出错')
            QMessageBox.warning(self, '警告', 'OpenSees终端无效！')
            self.ui.radioButton.setChecked(True)
            return
        os.system(f'{self.OS_terminal} {path1}/OS_terminal_test.tcl')
        if not os.path.exists(f'{path1}/ok.txt'):
            QMessageBox.warning(self, '警告', 'OpenSees终端无效！')
            os.remove(f"{path1}/OS_terminal_test.tcl")
            self.ui.radioButton.setChecked(True)
            return
        with open(f"{path1}/ok.txt", 'r') as f:
            ok_text = f.read()
        if 'ok' in ok_text:
            print('【MyWin, choose_OS_terminal】OpenSees求解器有效！\n', self.OS_terminal)
        else:
            QMessageBox.warning(self, '警告', 'OpenSees终端无效！')
            os.remove(f"{path1}/OS_terminal_test.tcl")
            self.ui.radioButton.setChecked(True)
            return
        os.remove(f"{path1}/OS_terminal_test.tcl")
        os.remove(f"{path1}/ok.txt")
        self.has_os_terminal = True

    @classmethod
    def build_tcl_file(cls,
            N: int, 
            m: list[float],
            mat_lib: list,
            story_mat: list,
            th_path: Path | str,
            SF: float,
            dt: float,
            mode_num: int,
            has_damping: bool,
            zeta_mode: tuple[int, int],
            zeta: tuple[float, float],
            setting: list,
            path: Path | str,
            gm_name: str,
            NPTS: int,
            print_result: bool=True
        ):
        """修改tcl文件"""
        with open('core/run_OS.tcl', 'r', encoding='utf=8') as f:
            text = f.read()
        text1 = str(int(N))
        text2 = ' '.join([str(i) for i in m])
        text3 = []
        for i in range(len(mat_lib)):
            mat = ' '.join([str(i) for i in mat_lib[i][3:]])
            mat = f'[list {mat}]'
            text3.append(mat)
        text3 = ' '.join(text3)
        text4 = []
        for i, mat in enumerate(story_mat):
            text4.append('[list ' + ' '.join([str(i) for i in mat]) + ']')
        text4 = ' '.join(text4)
        text5 = th_path
        text6 = str(float(SF))
        text7 = str(float(dt))
        text8 = str(int(mode_num))
        text8_1 = '1' if has_damping else '0'
        zeta_mode = [zeta_mode[0], zeta_mode[1]]
        text9 = ' '.join([str(i) for i in zeta_mode])
        zeta = [zeta[0], zeta[1]]
        text10 = ' '.join([str(i) for i in zeta])
        text11 = cls.setting1[setting[0]] + ' '
        text11 += cls.setting2[setting[1]] + ' '
        text11 += cls.setting3[setting[2]] + ' '
        text11 += cls.setting4[setting[3]] + ' '
        text11 += cls.setting5[setting[4]] + ' '
        text11 += cls.setting6[setting[5]] + ' '
        if setting[6]:
            text11 += str(setting[6]) + ' '
        else:
            text11 += '\"\" '
        if setting[7]:
            text11 += str(setting[7]) + ' '
        else:
            text11 += '\"\" '
        text11 += str(setting[8]) + ' '
        text11 += str(setting[9]) + ' '
        text11 += str(setting[10]) + ' '
        text11 += str(setting[11]) + ' '
        text11 += str(setting[12]) + ' '
        text11 += str(setting[13]) + ' '
        text11 += str(setting[14]) + ''
        text12 = path.replace('\\', '/')
        text13 = gm_name
        text14 = str(int(NPTS))
        text15 = str(float(cls.g))
        pattern = re.compile(r'(set N )\d+\n')
        text = pattern.sub(r'\g<1>' + text1 + r'\n', text)
        pattern = re.compile(r'(set m \[list )[1-9 ]+(\])')
        text = pattern.sub(r'\g<1>' + text2 + r'\2', text)
        pattern = re.compile(r'(set mat_lib \[list ).+(\]\n)')
        text = pattern.sub(r'\g<1>' + text3 + r'\2', text)
        pattern = re.compile(r'(set story_mat \[list ).+(\]\n)')
        text = pattern.sub(r'\g<1>' + text4 + r'\2', text)
        pattern = re.compile(r'(set th_path ").+(";)')
        text = pattern.sub(r'\g<1>' + text5 + r'\2', text)
        pattern = re.compile(r'(set SF )\d+(\n)')
        text = pattern.sub(r'\g<1>' + text6 + r'\2', text)
        pattern = re.compile(r'(set dt )[.0-9]+(\n)')
        text = pattern.sub(r'\g<1>' + text7 + r'\2', text)
        pattern = re.compile(r'(set mode_num )\d+(\n)')
        text = pattern.sub(r'\g<1>' + text8 + r'\2', text)
        pattern = re.compile(r'(set has_damping )\d+(\n)')
        text = pattern.sub(r'\g<1>' + text8_1 + r'\2', text)
        pattern = re.compile(r'(set zeta_mode \[list ).+(\]\n)')
        text = pattern.sub(r'\g<1>' + text9 + r'\2', text)
        pattern = re.compile(r'(set zeta \[list ).+(\]\n)')
        text = pattern.sub(r'\g<1>' + text10 + r'\2', text)
        pattern = re.compile(r'(set setting \[list ).+(\]\n)')
        text = pattern.sub(r'\g<1>' + text11 + r'\2', text)
        pattern = re.compile(r'(set path ").+("\n)')
        text = pattern.sub(r'\g<1>' + text12 + r'\2', text)
        pattern = re.compile(r'(set gm_name ").+("\n)')
        text = pattern.sub(r'\g<1>' + text13 + r'\2', text)
        pattern = re.compile(r'(set NPTS )\d+(\n)')
        text = pattern.sub(r'\g<1>' + text14 + r'\2', text)
        pattern = re.compile(r'(set g )[.0-9]+(\n)')
        text = pattern.sub(r'\g<1>' + text15 + r'\2', text)
        if print_result:
            text = re.sub('set print_results 0', 'set print_results 1', text)
        return text
    
    def clicked_build_tcl_file(self):
        if self.ready_to_run():
            th_path = 'xxx'
            zeta_mode = [self.ui.comboBox_3.currentIndex() + 1, self.ui.comboBox_4.currentIndex() + 1]
            zeta = [self.ui.lineEdit_3.text(), self.ui.lineEdit_3.text()]
            SF = 1
            text = MyWin.build_tcl_file(self.N, self.m, self.mat_lib, self.story_mat, th_path,
                                  SF, self.gm_dt[0], self.mode_num, self.has_damping, zeta_mode, zeta,
                                  self.setting, TEMP_PATH, self.gm_name[0], self.gm_NPTS[0])
        else:
            text = '模型未定义完全！'
        win = Win_tcl_file(text)
        win.exec_()

    # ----------------------------- tab 3 (post-processing) ------------------------

    def running_finished(self):
        print('【MyWin, running_finished】全部计算完成！')
        self.mode_results = core.ModeResults.from_file(self.mode_num, TEMP_PATH)
        self.all_resutls: list[core.Results] = []
        for i in range(self.gm_N):
            results = core.Results.from_file(self.gm_name[i], TEMP_PATH)
            self.all_resutls.append(results)
        self.result_exists = True
        self.update_result_combobox(self.ui.comboBox_5.currentIndex(), True)

    def update_result_combobox(self, idx=0, plot_curve=False):
        """更新结果combox的选项
            计算结果选择振型时，右侧列表框为振型
            计算结果选择其他项时，右侧列表框为地震动      
        """
        self.ui.comboBox_6.clear()
        self.ui.comboBox_7.clear()
        self.ui.comboBox_8.clear()
        item = self.ui.comboBox_5.itemText(idx)
        if item == '振型':
            self.ui.label_25.setText('振型：')
            self.ui.label_23.setEnabled(False)
            self.ui.comboBox_6.setEnabled(False)
            self.ui.label_24.setEnabled(False)
            self.ui.comboBox_7.setEnabled(False)
            for i in range(self.mode_num):
                self.ui.comboBox_8.addItem(f'{i+1}阶振型')
            if plot_curve:
                self.plot_results()
        elif item in ['相对位移', '相对速度', '相对加速度',
                      '绝对位移', '绝对速度', '绝对加速度', '楼层剪力']:
            self.ui.label_25.setText('地震动：')
            self.ui.label_23.setEnabled(True)
            self.ui.comboBox_6.setEnabled(True)
            self.ui.label_24.setEnabled(True)
            self.ui.comboBox_7.setEnabled(False)
            for i in range(self.gm_N):
                self.ui.comboBox_8.addItem(f'({i+1}) {self.gm_name[i]}')
            for i in range(self.N):
                self.ui.comboBox_6.addItem(str(i + 1))
            if plot_curve:
                self.plot_results()
        elif item in ['最大层间位移', '最大层间残余位移', '底部剪力',
                      '楼层剪力包络', '绝对加速度包络']:
            self.ui.label_25.setText('地震动：')
            self.ui.label_23.setEnabled(True)
            self.ui.comboBox_6.setEnabled(False)
            self.ui.label_24.setEnabled(True)
            self.ui.comboBox_7.setEnabled(False)
            for i in range(self.gm_N):
                self.ui.comboBox_8.addItem(f'({i+1}) {self.gm_name[i]}')
            for i in range(self.N):
                self.ui.comboBox_6.addItem(str(i + 1))
            if plot_curve:
                self.plot_results()
        elif item == '材料滞回曲线':
            self.ui.label_25.setText('地震动：')
            self.ui.label_23.setEnabled(True)
            self.ui.comboBox_6.setEnabled(True)
            self.ui.label_24.setEnabled(True)
            self.ui.comboBox_7.setEnabled(True)
            for i in range(self.gm_N):
                self.ui.comboBox_8.addItem(f'({i+1}) {self.gm_name[i]}')
            for i in range(self.N):
                self.ui.comboBox_6.addItem(str(i + 1))
            idx_story = self.ui.comboBox_6.currentIndex()
            self.update_hyeteretic_curve_list(idx_story)
            if plot_curve:
                self.plot_results()
        
    def update_hyeteretic_curve_list(self, idx_story):
        self.ui.comboBox_7.clear()
        if not self.ui.comboBox_5.currentText() == '材料滞回曲线':
            return
        if len(self.story_mat[idx_story]) == 1:
            # 该层只有一种材料
            mat_tag = self.story_mat[idx_story][0]
            mat_name = self.mat_lib[mat_tag - 1][0]
            self.ui.comboBox_7.addItem(mat_name)
        else:
            # 该层有多种材料
            for j, mat_tag in enumerate(self.story_mat[idx_story]):
                mat_name = self.mat_lib[mat_tag - 1][0]
                self.ui.comboBox_7.addItem(mat_name)
            else:
                self.ui.comboBox_7.addItem('并联材料')

    def plot_results(self, idx=None, skip=False):
        """当4个comboBox中任意一个的选项发生变化时触发"""
        if skip:
            print(skip)
            return
        if self.ui.comboBox_5.currentText() == '振型':
            x = list(range(0, self.N + 1, 1))
            y = self.mode_results(self.ui.comboBox_8.currentIndex() + 1)
            y = np.insert(y, 0, 0)
            self.plot_result_mode(x, y, f'第{self.ui.comboBox_8.currentIndex()+1}阶振型')
            self.update_graph_data(x, y, '振型', '楼层', '位移')
        else:
            story_id = self.ui.comboBox_6.currentIndex() + 1
            gm_idx = self.ui.comboBox_8.currentIndex()
            gm_name = self.gm_name[gm_idx]
            results = self.all_resutls[gm_idx]
            t = results.t
        if self.ui.comboBox_5.currentText() == '相对位移':
            ru = results.ru[:, story_id - 1]
            case_ = f'{gm_name}第{story_id}层相对位移'
            self.plot_result_th(t, ru, 't [s]', '相对位移 [mm]', case_)
            self.update_graph_data(t, ru, case_, 't [s]', '相对位移 [mm]')
        elif self.ui.comboBox_5.currentText() == '相对速度':
            rv = results.rv[:, story_id - 1]
            case_ = f'{gm_name}第{story_id}层相对速度'
            self.plot_result_th(t, rv, 't [s]', '相对速度 [mm/s]', case_)
            self.update_graph_data(t, rv, case_, 't [s]', '相对速度 [mm/s]')
        elif self.ui.comboBox_5.currentText() == '相对加速度':
            ra = results.ra[:, story_id - 1] / self.g
            case_ = f'{gm_name}第{story_id}层相对加速度'
            self.plot_result_th(t, ra, 't [s]', '相对加速度 [g]', case_)
            self.update_graph_data(t, ra, case_, 't [s]', '相对加速度 [g]')
        elif self.ui.comboBox_5.currentText() == '最大层间位移':
            x_story = list(range(0, self.N + 1, 1))
            ru = results.ru
            interstory_ru = np.diff(ru, axis=1)
            interstory_ru = np.column_stack((ru[:, 0], interstory_ru))
            max_IDR = np.amax(abs(interstory_ru), axis=0)
            max_IDR = np.insert(max_IDR, 0, 0)
            case_ = f'{gm_name}最大层间位移'
            self.plot_result_th(x_story, max_IDR, '楼层', '最大层间位移 [mm]', case_, True)
            self.update_graph_data(x_story, max_IDR, case_, '楼层', '最大层间位移 [mm]')
        elif self.ui.comboBox_5.currentText() == '最大层间残余位移':
            x_story = list(range(0, self.N + 1, 1))
            resu = results.resu
            RIDR = np.diff(resu)
            RIDR = np.insert(RIDR, 0, resu[0])
            RIDR = np.insert(RIDR, 0, 0)
            case_ = f'{gm_name}最大层间残余位移'
            self.plot_result_th(x_story, RIDR, '楼层', '最大层间残余位移 [mm]', case_, True)
            self.update_graph_data(x_story, RIDR, case_, '楼层', '最大层间残余位移 [mm]')
        elif self.ui.comboBox_5.currentText() == '绝对位移':
            au = results.au[:, story_id - 1]
            case_ = f'{gm_name}第{story_id}层绝对位移'
            self.plot_result_th(t, au, 't [s]', '绝对位移 [mm]', case_)
            self.update_graph_data(t, au, case_, 't [s]', '绝对位移 [mm]')
        elif self.ui.comboBox_5.currentText() == '绝对速度':
            av = results.av[:, story_id - 1]
            case_ = f'{gm_name}第{story_id}层绝对速度'
            self.plot_result_th(t, av, 't [s]', '绝对速度 [mm/s]', case_)
            self.update_graph_data(t, av, case_, 't [s]', '绝对速度 [mm/s]')
        elif self.ui.comboBox_5.currentText() == '绝对加速度':
            aa = results.aa[:, story_id - 1] / self.g
            case_ = f'{gm_name}第{story_id}层绝对加速度'
            self.plot_result_th(t, aa, 't [s]', '绝对加速度 [g]', case_)
            self.update_graph_data(t, aa, case_, 't [s]', '绝对加速度 [g]')
        elif self.ui.comboBox_5.currentText() == '楼层剪力':
            story_idx = self.ui.comboBox_6.currentIndex()
            mat_idx = self.story_mat.copy()
            n = 0
            stressStrain = results.mat
            for i, sub_list in enumerate(mat_idx):
                for j, _ in enumerate(sub_list):
                    mat_idx[i][j] = n
                    n += 1
            if len(self.story_mat[story_idx]) == 1:
                # 该层只有一个材料
                col_idx = mat_idx[story_idx][0] * 2  # 材料结果数据的列数索引
                F = stressStrain[:, col_idx]
            else:
                # 该层有多种材料
                F = np.zeros(len(t))
                for i in range(len(mat_idx[story_idx])):
                    # 遍历每一种单独材料并叠加
                    col_idx = mat_idx[story_idx][i] * 2
                    F_temp = stressStrain[:, col_idx]
                    F += F_temp
            case_ = f'{gm_name}第{story_idx+1}层层间剪力'
            self.plot_result_th(t, F / 1000, 't [s]', '力 [kN]', case_)
            self.update_graph_data(t, F / 1000, case_, 't [s]', '力 [kN]')
        elif self.ui.comboBox_5.currentText() == '底部剪力':
            base_V = results.base_V / 1000
            case_ = f'{gm_name}底部剪力'
            self.plot_result_th(t, base_V, 't [s]', '底部剪力 [kN]', case_)
            self.update_graph_data(t, base_V, case_, 't [s]', '底部剪力 [kN]')
        elif self.ui.comboBox_5.currentText() == '材料滞回曲线':
            story_idx = self.ui.comboBox_6.currentIndex()
            stressStrain = results.mat
            mat_idx: list[list[int]] = []
            n = 0
            for i, sub_list in enumerate(self.story_mat):
                mat_idx.append([])
                for j, _ in enumerate(sub_list):
                    mat_idx[i].append(n)
                    n += 1
            if len(self.story_mat[story_idx]) == 1:
                # 该层只有一个材料
                col_idx = mat_idx[story_idx][0] * 2  # 材料结果数据的列数索引
                u = stressStrain[:, col_idx + 1]
                F = stressStrain[:, col_idx]
            else:
                # 该层有多种材料
                if self.ui.comboBox_7.currentIndex() != len(mat_idx[story_idx]):
                    # 选择单独材料
                    col_idx = mat_idx[story_idx][self.ui.comboBox_7.currentIndex()] * 2
                    u = stressStrain[:, col_idx + 1]
                    F = stressStrain[:, col_idx]
                else:
                    # 选中并联材料
                    u, F = np.zeros(len(t)), np.zeros(len(t))
                    for i in range(len(mat_idx[story_idx])):
                        # 遍历每一种单独材料并叠加
                        col_idx = mat_idx[story_idx][i] * 2
                        u_temp = stressStrain[:, col_idx + 1]
                        F_temp = stressStrain[:, col_idx]
                        u = u_temp
                        F += F_temp
            case_ = f'{gm_name}第{story_idx+1}层材料滞回曲线'
            self.plot_result_th(u, F / 1000, '位移 [mm]', '力 [kN]', case_)
            self.update_graph_data(u, F / 1000, case_, '位移 [mm]', '力 [kN]')
        elif self.ui.comboBox_5.currentText() == '楼层剪力包络':
            x_story = list(range(0, self.N + 1, 1))[1:]
            story_idx = self.ui.comboBox_6.currentIndex()
            stressStrain = results.mat
            mat_idx = self.story_mat.copy()
            n = 0
            for i, sub_list in enumerate(mat_idx):
                for j, _ in enumerate(sub_list):
                    mat_idx[i][j] = n
                    n += 1
            F = np.zeros((len(t), self.N))
            for story_idx in range(self.N):
                if len(self.story_mat[story_idx]) == 1:
                    # 该层只有一个材料
                    col_idx = mat_idx[story_idx][0] * 2  # 材料结果数据的列数索引
                    F_i = stressStrain[:, col_idx]
                else:
                    # 该层有多种材料
                    F_i = np.zeros(len(t))
                    for i in range(len(mat_idx[story_idx])):
                        # 遍历每一种单独材料并叠加
                        col_idx = mat_idx[story_idx][i] * 2
                        F_temp = stressStrain[:, col_idx]
                        F_i += F_temp
                F[:, story_idx] = F_i
            F = np.amax(abs(F), axis=0) / 1000
            case_ = f'{gm_name}第{story_idx+1}层层间剪力包络'
            self.plot_result_th(x_story, F, '楼层', '最大剪力 [kN]', case_, True)
            self.update_graph_data(x_story, F, case_, '楼层', '最大剪力 [kN]')
        elif self.ui.comboBox_5.currentText() == '绝对加速度包络':
            x_story = list(range(0, self.N + 1, 1))[1:]
            aa = results.aa / self.g
            max_aa = np.max(abs(aa), axis=0)
            case_ = f'{gm_name}绝对加速度包络'
            self.plot_result_th(x_story, max_aa, 't [s]', '绝对加速度包络 [g]', case_, True)
            self.update_graph_data(x_story, max_aa, case_, 't [s]', '绝对加速度包络 [g]')
        self.display_period()

    def plot_result_th(self, x, y, x_label, y_label, case_, plot_scatter=False):
        print('【MyWin, plot_result_mode】绘制时程曲线 - ' + case_)
        self.pg3.clear()
        line = pg.PlotCurveItem(x, y, pen=self.pen1)
        self.pg3.addItem(line)
        if plot_scatter:
            line1 = pg.ScatterPlotItem(x, y, size=12, brush=(68, 114, 196))
            self.pg3.addItem(line1)
        self.pg3.setLabel(axis='left', text=y_label)
        self.pg3.setLabel(axis='bottom', text=x_label)
        self.pg3.autoRange()
        
    def plot_result_mode(self, x, y, case_=None):
        print('【MyWin, plot_result_mode】绘制振型 - ' + case_)
        self.pg3.clear()
        line1 = pg.PlotCurveItem(x, y, pen=self.pen1)
        line2 = pg.ScatterPlotItem(x, y, size=12, brush=(68, 114, 196))
        self.pg3.addItem(line1)
        self.pg3.addItem(line2)
        self.pg3.setLabel(axis='left', text='位移 [mm]')
        self.pg3.setLabel(axis='bottom', text='楼层')
        self.pg3.autoRange()

    def display_period(self):
        if self.ui.comboBox_5.currentText() == '振型':
            n = self.ui.comboBox_8.currentIndex() + 1
            T = self.mode_results.T[n - 1]
            self.export_set_text(f'T{n} = {T:.4f} s')
        else:
            self.export_set_text('')

    def update_graph_data(self, x, y, title, x_title, y_title):
        """绘图数据"""
        self.current_plot_data = (x, y, title, x_title, y_title)

    def show_graph_data(self):
        win = Win_data(self)
        win.exec_()

    def export_data(self):
        if not self.result_exists:
            QMessageBox.warning(self, '警告', '无数据！')
            return
        self.ui.pushButton_22.setEnabled(False)
        self.ui.pushButton_19.setEnabled(False)
        self.ui.pushButton_22.setText('正在导出...')
        win = Win_export(self)
        win.exec_()
        if self.export_type == 'xlsx':
            self.export_type = None
            excel_path, _ = QFileDialog.getSaveFileName(self, "保存文件", "结果统计.xlsx", "Excel Files (*.xlsx)")
            if not excel_path:
                self.ui.pushButton_22.setEnabled(True)
                self.ui.pushButton_19.setEnabled(True)
                self.ui.pushButton_22.setText('导出所有数据')
                return
            print(f'【MyWin, export_data】导出路径：{excel_path}')
            # self.export_data_xlsx(excel_path)
            self.thread_export_data = Thread_export_data(self, 'xlsx', excel_path)
            self.thread_export_data.signal_finished.connect(self.export_finished)
            self.thread_export_data.signal_msg.connect(self.export_message)
            self.thread_export_data.signal_info.connect(self.export_set_text)
            self.thread_export_data.start()
        elif self.export_type == 'txt':
            self.export_type = None
            txt_path = QFileDialog.getExistingDirectory(self, '选择保存路径')
            if not txt_path:
                self.ui.pushButton_22.setEnabled(True)
                self.ui.pushButton_19.setEnabled(True)
                self.ui.pushButton_22.setText('导出所有数据')
                return
            print(f'【MyWin, export_data】导出路径：{txt_path}')
            # self.export_data_txt(txt_path)
            self.thread_export_data = Thread_export_data(self, 'txt', txt_path)
            self.thread_export_data.signal_finished.connect(self.export_finished)
            self.thread_export_data.signal_msg.connect(self.export_message)
            self.thread_export_data.signal_info.connect(self.export_set_text)
            self.thread_export_data.start()
        else:
            self.export_finished()
        # self.ui.pushButton_22.setEnabled(True)
        # self.ui.pushButton_19.setEnabled(True)
        # self.ui.pushButton_22.setText('导出所有数据')

    def export_finished(self):
        self.ui.pushButton_22.setEnabled(True)
        self.ui.pushButton_19.setEnabled(True)
        self.ui.pushButton_22.setText('导出所有数据')
        self.statusBar_label_right.setText('')

    def export_message(self, list_):
        msg_type = list_[0]
        str_ = list_[1]
        if msg_type == 'information':
            QMessageBox.information(self, '提示', str_)
        elif msg_type == 'critical':
            QMessageBox.critical(self, '错误', str_)
        elif msg_type == 'warning':
            QMessageBox.warning(self, '警告', str_)

    def export_set_text(self, text):
        self.statusBar_label_right.setText(text)

    @staticmethod
    def is_iterable(obj):
        try:
            iter(obj)
            return True
        except TypeError:
            return False

    @staticmethod
    def write_to_excel(ws, data: np.ndarray, row_start: int, col_start: int, fmt: int=7):
        """写入一维或二维数组到excel

        Args:
            ws (_type_): _description_
            data (np.ndarray): 一维或二维ndarray数组
            row_start (int): 起始行序号
            col_start (int): 起始列序号
            fmt (int, optional): 保留小数位数，默认7
        """
        # 写入一维或二维数组到excel
        for i, rows in enumerate(data):
            if MyWin.is_iterable(rows):
                for j, val in enumerate(rows):
                    val = round(val, fmt)
                    ws.cell(row=row_start+i, column=col_start+j, value=val)
            else:
                ws.cell(row=row_start+i, column=col_start, value=rows)

    def set_ws_center(self, ws):
        if self.gm_N * self.N > 8:
            return  # 如果地震动和层数过多，则不进行居中以节省时间
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = px.styles.Alignment(horizontal='center', vertical='center')


    # ----------------------------- Menu ------------------------------------------

    def setting_clicked(self):
        win = Win_setting(self)
        win.exec_()

    def closeEvent(self, event):
        print('【MyWin, closeEvent】退出')
        if os.path.exists(f'{TEMP_PATH}/temp_NLMDOF_results'):
            rmtree(f'{TEMP_PATH}/temp_NLMDOF_results')
        super().closeEvent(event)

    def open_win_about(self):
        win = Win_about()
        win.exec_()


class Win_importGM(QDialog):
    """导入地震动窗口"""
    def __init__(self, main: MyWin):
        super().__init__()
        self.ui = Ui_win_importGM()
        self.main = main
        self.ui.setupUi(self)
        self.init_ui()
    
    def init_ui(self):
        self.ui.checkBox.toggled.connect(lambda: self.ui.lineEdit_2.setEnabled(self.ui.checkBox.isChecked()))
        self.ui.radioButton_5.toggled.connect(lambda: self.ui.lineEdit.setEnabled(self.ui.radioButton_5.isChecked()))
        self.ui.pushButton.clicked.connect(self.choose_gm)
        self.ui.lineEdit_3.editingFinished.connect(self.change_dt)
        self.ui.lineEdit.editingFinished.connect(self.change_PGA)
        self.ui.lineEdit_4.editingFinished.connect(self.change_SF)
        self.ui.radioButton_6.toggled.connect(lambda: self.ui.lineEdit_4.setEnabled(self.ui.radioButton_6.isChecked()))
        validator = QDoubleValidator(0, 10000, 7)
        validator.setNotation(QDoubleValidator.StandardNotation)
        self.ui.lineEdit_3.setValidator(validator)
        self.ui.lineEdit.setValidator(validator)
        self.ui.lineEdit_4.setValidator(validator)
        self.ui.radioButton_2.toggled.connect(lambda: self.ui.lineEdit_3.setEnabled(not self.ui.lineEdit_3.isEnabled()))
        self.ui.pushButton_2.clicked.connect(self.choose_records_file)

    def choose_gm(self):
        skip_rows = int(self.ui.lineEdit_2.text())
        dt = float(self.ui.lineEdit_3.text())
        paths = QFileDialog.getOpenFileNames(self, '导入地震动')[0]
        if not paths:
            print('【Win_importGM, choose_gm】没有导入地震动。')
            return 0
        for path in paths:
            try:
                data = np.loadtxt(path, dtype=str, encoding='utf-8', skiprows=skip_rows)
                if len(data) == 0:
                    QMessageBox.warning(self, '警告', '数据为空！')
                    return 0
                data = data.astype(float)
                if self.ui.radioButton.isChecked():
                    # 单列加速度
                    th = data
                else:
                    # 时间&加速度
                    t = data[:, 0]
                    th = data[:, 1]
                    dt = t[1] - t[0]
                    for i in range(len(t) - 1):
                        if t[i + 1] - t[i] <= 0:
                            QMessageBox.warning(self, '警告', '时间序列不是单调递增的！')
                            return 0
                    for ti in t:
                        if ti < 0:
                            QMessageBox.warning(self, '警告', '时间序列存在负数！')
                            return 0   
            except:
                QMessageBox.warning(self, '警告', f'"{path}"无法读取！')
                self.close_win()
                return 0
            # 缩放选项
            if self.ui.radioButton_4.isChecked():  # 归一化
                if max(abs(th)) == 0:
                    QMessageBox.warning(self, '警告', '数据均为0！')
                    self.close_win()
                    return 0
                th = th / max(abs(th))
            elif self.ui.radioButton_5.isChecked():  # 指定PGA
                if max(abs(th)) == 0:
                    QMessageBox.warning(self, '警告', f'数据均为0！')
                    self.close_win()
                    return 0
                PGA = float(self.ui.lineEdit.text())
                th = th / max(abs(th)) * PGA
            elif self.ui.radioButton_6.isChecked():  # 指定缩放系数
                th *= float(self.ui.lineEdit_4.text())
            self.main.gm.append(th)
            N = len(self.main.gm[-1]) - 1
            self.main.gm_NPTS.append(N)
            self.main.gm_duration.append(self.main.gm_NPTS[-1] * dt)
            if self.ui.radioButton.isChecked():
                # self.main.gm_t.append(np.arange(0, self.main.gm_duration[-1] + dt, dt))
                self.main.gm_t.append(np.linspace(0, self.main.gm_NPTS[-1] * dt, N + 1))
            else:
                self.main.gm_t.append(t)
            self.main.gm_N += 1
            self.main.gm_dt.append(dt)
            self.main.gm_name.append(os.path.basename(path).split('.')[0])
            self.main.gm_unit.append('g')
            self.main.gm_PGA.append(max(abs(th)))
        self.close_win()

    def choose_records_file(self):
        """通过.records文件导入"""
        path = Path(QFileDialog.getOpenFileName(self, '选择.records文件')[0])
        if not path:
            print('【Win_importGM, choose_gm】没有导入地震动。')
            return 0
        if not path.suffix == '.records':
            QMessageBox.warning(self, '错误', '请选择.records文件！')
            return 0
        if not path.exists():
            QMessageBox.warning(self, '错误', '文件不存在！')
            return 0
        try:
            with open(path, 'rb') as f:
                records: su.Records = dill.load(f)
        except Exception as e:
            QMessageBox.warning(self, '错误', '文件格式错误！')
            raise e
            return 0
        if self.ui.checkBox_2.isChecked():
            gen = records.get_scaled_records()
        else:
            gen = records.get_unscaled_records()
        print(records.get_record_name())
        for i, (th, dt) in enumerate(gen):
            t = np.arange(0, len(th) * dt, dt)
            self.main.gm.append(th)
            self.main.gm_NPTS.append(len(th))
            self.main.gm_duration.append((self.main.gm_NPTS[-1] - 1) * dt)
            self.main.gm_t.append(t)
            self.main.gm_N += 1
            self.main.gm_dt.append(dt)
            name = records.get_record_name()[i].replace('/', '_')
            name = name.replace('\\', '_')
            self.main.gm_name.append(name)
            self.main.gm_unit.append('g')
            self.main.gm_PGA.append(max(abs(th)))
        self.close_win()

    def change_dt(self):
        if float(self.ui.lineEdit_3.text()) == 0:
            self.ui.lineEdit_3.setText('0.00001')

    def change_PGA(self):
        if float(self.ui.lineEdit.text()) == 0:
            self.ui.lineEdit.setText('0.00001')

    def change_SF(self):
        if float(self.ui.lineEdit_4.text()) == 0:
            self.ui.lineEdit_4.setText('0.00001')

    def close_win(self):
        self.accept()
        self.main.gm_list_update()


class Win_importGM1(QDialog):
    """导入地震动窗口"""
    gm_name = {0: 'ChiChi', 1: 'Friuli', 2: 'Hollister', 3: 'Imperial_Valley', 4: 'Kobe',\
               5: 'Kocaeli', 6: 'Landers', 7: 'Loma_Prieta', 8: 'Northridge', 9: 'Trinidad'}

    def __init__(self, main: MyWin):
        super().__init__()
        self.ui = Ui_win_importGM1()
        self.main = main
        self.ui.setupUi(self)
        self.init_ui()

    def init_ui(self):
        self.ui.pushButton_2.clicked.connect(lambda: self.accept())
        self.ui.pushButton.clicked.connect(self.choose_gm1)

    def choose_gm1(self):
        """选择常用地震动"""
        if not self.ui.listWidget.selectedItems():
            print('【Win_importGM1, choose_gm1】没有选择项目')
            self.accept()
            return
        idx = self.ui.listWidget.currentRow()
        print(f'【Win_importGM1, choose_gm1】选中:{idx}')
        # 使用__file__相对路径以确保打包后可找到数据文件
        data = np.loadtxt(Path(__file__).parent.parent / f'data/{Win_importGM1.gm_name[idx]}.dat', dtype=float)
        t, th = data[:, 0], data[:, 1]
        dt = t[1] - t[0]
        self.main.gm.append(th)
        self.main.gm_name.append(Win_importGM1.gm_name[idx])
        self.main.gm_N += 1
        self.main.gm_dt.append(dt)
        self.main.gm_NPTS.append(len(th) - 1)
        self.main.gm_t.append(t)
        self.main.gm_duration.append(t[-1])
        self.main.gm_unit.append('g')
        self.main.gm_PGA.append(max(abs(th)))
        self.main.gm_list_update()
        self.accept()


class Win_mass(QDialog):
    """定义质量窗口"""
    def __init__(self, main: MyWin):
        super().__init__()
        self.ui = Ui_win_mass()
        self.main = main
        self.ui.setupUi(self)
        self.init_ui()

    def init_ui(self):
        self.ui.pushButton.clicked.connect(self.ok)
        self.ui.pushButton_2.clicked.connect(self.cancel)
        self.ui.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.ui.tableWidget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.ui.tableWidget.customContextMenuRequested.connect(self.showContextMenu)
        if self.main.N:
            self.ui.tableWidget.setRowCount(self.main.N)
            for i in range(self.main.N):
                self.ui.tableWidget.setItem(i, 0, QTableWidgetItem(f'm{i+1}'))
                item = self.ui.tableWidget.item(i, 0)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            if self.main.m:
                for i, mi in enumerate(self.main.m):
                    self.ui.tableWidget.setItem(i, 1, QTableWidgetItem(str(mi)))
            for r in range(self.ui.tableWidget.rowCount()):
                for c in range(self.ui.tableWidget.columnCount()):
                    if self.ui.tableWidget.item(r, c):
                        self.ui.tableWidget.item(r, c).setTextAlignment(0x0004 | 0x0080)

    def ok(self):
        row = self.ui.tableWidget.rowCount()
        if not row:
            self.accept()
            return
        try:
            m = []
            for i in range(row):
                m.append(float(self.ui.tableWidget.item(i, 1).text()))
            else:
                self.main.m = m
            self.accept()
        except:
            QMessageBox.warning(self, '警告', '无法读取表格数据！')

    def cancel(self):
        self.accept()

    def showContextMenu(self, pos):
        """定义上下文菜单"""
        context_menu = QMenu(self)
        copy_action = context_menu.addAction("复制")
        paste_action = context_menu.addAction("粘贴")
        menu_size = context_menu.sizeHint()  # 获取菜单的大小
        global_pos = self.ui.tableWidget.mapToGlobal(pos)  # 获取全局坐标
        adjusted_pos = global_pos + QPoint(0, menu_size.height() // 2)  # 调整位置以使其出现在鼠标指针的右下方
        action = context_menu.exec_(adjusted_pos)
        if action == copy_action:
            self.copy_selected_cells()
        elif action == paste_action:
            self.paste_to_cells()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_C and event.modifiers() == Qt.ControlModifier:
            self.copy_selected_cells()
        elif event.key() == Qt.Key_V and event.modifiers() == Qt.ControlModifier:
            self.paste_to_cells()
        else:
            super().keyPressEvent(event)

    def copy_selected_cells(self):
        selected_items = self.ui.tableWidget.selectedItems()
        if not selected_items:
            return
        text = ''
        row_start = selected_items[0].row()
        row_end = selected_items[-1].row()
        col_start = selected_items[0].column()
        col_end = selected_items[-1].column()
        for r in range(row_start, row_end + 1):
            row_data = []
            for c in range(col_start, col_end + 1):
                item = self.ui.tableWidget.item(r, c)
                if item and item.text():
                    row_data.append(item.text())
            text += "\t".join(row_data) + "\n"
        clipboard = QApplication.clipboard()
        clipboard.setText(text)

    def paste_to_cells(self):
        current_row = self.ui.tableWidget.currentRow()
        current_col = self.ui.tableWidget.currentColumn()
        max_row = self.ui.tableWidget.rowCount()
        max_col = self.ui.tableWidget.columnCount()
        clipboard = QApplication.clipboard()
        text = clipboard.text()
        if not text:
            return
        rows = text.strip().split('\n')
        for i, row in enumerate(rows):
            cols = row.split('\t')
            for j, value in enumerate(cols):
                if i + current_row >= max_row or j + current_col >= max_col:
                    continue
                if j + current_col == 0:
                    continue
                item = QTableWidgetItem(value)
                self.ui.tableWidget.setItem(i + current_row, j + current_col, item)
                self.ui.tableWidget.item(i + current_row, j + current_col).setTextAlignment(0x0004 | 0x0080)



class Win_mat(QDialog):
    """定义材料窗口"""
    def __init__(self, main: MyWin, mat_idx, mod_idx=None):
        super().__init__()
        self.ui = Ui_win_mat()
        self.main = main
        self.mat_idx = mat_idx
        self.mod_idx = mod_idx  # 修改模式下材料的idx
        self.ui.setupUi(self)
        self.init_ui()

    def init_ui(self):
        mat = ['线弹性模型', '双线性模型', 'Wen模型', '理想弹塑性模型', '黏性模型'][self.mat_idx]
        self.ui.label.setText(f'材料类型：{mat}')
        self.ui.pushButton_2.clicked.connect(lambda: self.accept())
        self.ui.pushButton.clicked.connect(self.ok)
        self.ui.lineEdit_5.setPlaceholderText('选填')
        validator = QDoubleValidator(0, 1e15, 7)
        validator.setNotation(QDoubleValidator.StandardNotation)
        self.ui.lineEdit.setValidator(validator)
        self.ui.lineEdit_2.setValidator(validator)
        self.ui.lineEdit_3.setValidator(validator)
        self.ui.lineEdit_4.setValidator(validator)
        if type(self.mod_idx) is int:
            print('【Win_mat, init_ui】正在修改')
            if self.main.mat_lib[self.mod_idx][2] == 1:
                print('【Win_mat, init_ui】有定义备注名')
                self.ui.lineEdit_5.setText(self.main.mat_lib[self.mod_idx][0])   # 如果定义材料时有填备注名，则显示备注名
            else:
                print('【Win_mat, init_ui】没有定义备注名')
        if self.mat_idx == 0:
            self.ui.label_2.setText('刚度（E [N/mm]）：')
            self.ui.label_3.clear()
            self.ui.label_4.clear()
            self.ui.label_5.clear()
            self.ui.lineEdit_2.hide()
            self.ui.lineEdit_3.hide()
            self.ui.lineEdit_4.hide()
            if type(self.mod_idx) is int:
                self.ui.lineEdit.setText(str(self.main.mat_lib[self.mod_idx][5]))
        elif self.mat_idx == 1:
            self.ui.label_2.setText('屈服力（Fy [N]）：')
            self.ui.label_3.setText('初始刚度（E [N/mm]）：')
            self.ui.label_4.setText('屈服后刚度比（α）：')
            self.ui.label_5.clear()
            self.ui.lineEdit_4.hide()
            if type(self.mod_idx) is int:
                self.ui.lineEdit.setText(str(self.main.mat_lib[self.mod_idx][5]))
                self.ui.lineEdit_2.setText(str(self.main.mat_lib[self.mod_idx][6]))
                self.ui.lineEdit_3.setText(str(self.main.mat_lib[self.mod_idx][7]))
        elif self.mat_idx == 2:
            self.ui.label_2.setText('屈服力（Fy [N]）：')
            self.ui.label_3.setText('屈服位移（uy [mm]）：')
            self.ui.label_4.setText('屈服后刚度比（α）：') 
            self.ui.label_5.setText('屈服指数（n）：')
            if type(self.mod_idx) is int:
                self.ui.lineEdit.setText(str(self.main.mat_lib[self.mod_idx][5]))
                self.ui.lineEdit_2.setText(str(self.main.mat_lib[self.mod_idx][6]))
                self.ui.lineEdit_3.setText(str(self.main.mat_lib[self.mod_idx][7]))
                self.ui.lineEdit_4.setText(str(self.main.mat_lib[self.mod_idx][8]))
        elif self.mat_idx == 3:
            self.ui.label_2.setText('屈服力（Fy [n]）：')
            self.ui.label_3.setText('初始刚度（E [N/mm]）：')
            self.ui.label_4.clear()
            self.ui.label_5.clear()
            self.ui.lineEdit_3.hide()
            self.ui.lineEdit_4.hide()
            if type(self.mod_idx) is int:
                self.ui.lineEdit.setText(str(self.main.mat_lib[self.mod_idx][5]))
                self.ui.lineEdit_2.setText(str(self.main.mat_lib[self.mod_idx][6]))
        elif self.mat_idx == 4:
            self.ui.label_2.setText('阻尼系数（C [N/(mm/s)^α]）：')
            self.ui.label_3.setText('阻尼指数（α）：')
            self.ui.label_4.clear()
            self.ui.label_5.clear()
            self.ui.lineEdit_3.hide()
            self.ui.lineEdit_4.hide()
            if type(self.mod_idx) is int:
                self.ui.lineEdit.setText(str(self.main.mat_lib[self.mod_idx][5]))
                self.ui.lineEdit_2.setText(str(self.main.mat_lib[self.mod_idx][6]))

    def ok(self):
        p1 = self.ui.lineEdit.text()
        p2 = self.ui.lineEdit_2.text()
        p3 = self.ui.lineEdit_3.text()
        p4 = self.ui.lineEdit_4.text()
        commentName = self.ui.lineEdit_5.text()
        para = []
        have_name = 0  # 是否有备注名，0或1
        if self.mat_idx == 0:  # 弹性
            if not commentName:
                commentName = f'弹性(E={p1})'
            else:
                have_name = 1
            if not p1:
                QMessageBox.warning(self, '警告', '存在参数未定义！')
                return
            else:
                para = [commentName, 0, have_name, 'Elastic', int(self.main.mat_N + 1), p1]
        elif self.mat_idx == 1:  # 双线性
            if not commentName:
                commentName = f'双线性(Fy={p1},E={p2},α={p3})'
            else:
                have_name = 1
            if not all([p1, p2, p3]):
                QMessageBox.warning(self, '警告', '存在参数未定义！')
                return
            else:
                para = [commentName, 1, have_name, 'Steel01', int(self.main.mat_N + 1), p1, p2, p3]
        elif self.mat_idx == 2:  # Wen
            if not commentName:
                commentName = f'Wen(Fy={p1},uy={p2},α={p3},n={p4})'
            else:
                have_name = 1
            if not all([p1, p2, p3, p4]):
                QMessageBox.warning(self, '警告', '存在参数未定义！')
                return
            else:
                para = [commentName, 2, have_name, 'BoucWen', int(self.main.mat_N + 1), p1, p2, p3, p4]
        elif self.mat_idx == 3:  # 理想弹塑性
            if not commentName:
                commentName = f'理想弹塑性(Fy={p1},E={p2})'
            else:
                have_name = 1
            if not (p1 and p2):
                QMessageBox.warning(self, '警告', '存在参数未定义！')
                return
            else:
                para = [commentName, 3, have_name, 'Steel01', int(self.main.mat_N + 1), p1, p2, 0]
        elif self.mat_idx == 4:  # 黏性
            if not commentName:
                commentName = f'黏性(E={p1})'
            else:
                have_name = 1
            if not (p1 and p2):
                QMessageBox.warning(self, '警告', '存在参数未定义！')
                return
            else:
                para = [commentName, 4, have_name, 'Viscous', int(self.main.mat_N + 1), p1, p2]
        for i in range(5, len(para)):
            para[i] = float(para[i])
        if type(self.mod_idx) is int:
            # 修改模式
            para[4] = self.main.mat_lib[self.mod_idx][4]
            self.main.ui.listWidget_2.item(self.mod_idx).setText(f'({int(para[4])}) {para[0]}')
            self.main.mat_lib[self.mod_idx] = para
            print('【Win_mat, ok】已修改：', para)
        else:
            # 添加模式
            self.main.ui.listWidget_2.addItem(f'({int(para[4])}) {para[0]}')
            self.main.mat_lib.append(para)
            self.main.mat_N += 1
            print('【Win_mat, ok】已定义：', para)
        self.accept()
        self.main.update_conbeBox_mat()


class Win_OSmat(QDialog):
    def __init__(self, main: MyWin, mod_idx=None):
        super().__init__()
        self.ui = Ui_win_OSmat()
        self.main = main
        self.ui.setupUi(self)
        self.mod_idx = mod_idx
        self.init_ui()
        self.ui_mod()

    def init_ui(self):
        self.ui.tableWidget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.ui.tableWidget.customContextMenuRequested.connect(self.showContextMenu)
        self.ui.pushButton_2.clicked.connect(self.accept)
        self.ui.spinBox.textChanged.connect(self.set_line_num)
        self.ui.lineEdit.setPlaceholderText('选填')
        self.ui.pushButton.clicked.connect(self.ok)
        item = self.ui.tableWidget.item(1, 0)
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        self.set_table_center()
        self.ui.tableWidget.cellChanged.connect(self.set_table_center)

    def set_table_center(self):
        """设置表格居中"""
        pass
        # for r in range(self.ui.tableWidget.rowCount()):
        #     for c in range(self.ui.tableWidget.columnCount()):
        #         if self.ui.tableWidget.item(r, c):
        #             self.ui.tableWidget.item(r, c).setTextAlignment(0x0004 | 0x0080)

    def ui_mod(self):
        """修改模式"""
        if type(self.mod_idx) is not int:
            return
        mat = self.main.mat_lib[self.mod_idx]
        print('【Win_OSmat, ui_mod】正在修改：', mat)
        self.ui.spinBox.setValue(len(mat) - 3)
        self.ui.lineEdit.setText(mat[0])
        self.ui.tableWidget.setItem(0, 0, QTableWidgetItem(str(mat[3])))
        self.ui.tableWidget.setItem(1, 0, QTableWidgetItem('tag'))
        for i in range(5, len(mat)):
            self.ui.tableWidget.setItem(i - 3, 0, QTableWidgetItem(str(mat[i])))
        item = self.ui.tableWidget.item(1, 0)
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        self.set_table_center()

    def set_line_num(self):
        new_count = int(self.ui.spinBox.text())
        self.ui.tableWidget.setRowCount(new_count)
        self.set_table_center()

    def check(self):
        """检查是否所有数据都已经输入"""
        count = self.ui.tableWidget.rowCount()
        if not self.ui.tableWidget.item(0, 0).text():
            return False
        for i in range(3, count):
            if not self.ui.tableWidget.item(i, 0).text():
                return False
        return True
    
    def ok(self):
        if not self.check():
            QMessageBox.warning(self, '警告', '存在参数未定义！')
            return
        mat = [None, None, None, None, None]
        mat_type = self.ui.tableWidget.item(0, 0).text()
        mat_para = []
        for i in range(2, self.ui.tableWidget.rowCount()):
            para = self.ui.tableWidget.item(i, 0).text()
            try:
                para = float(para)
            except:
                pass
            mat_para.append(para)
        if not self.ui.lineEdit.text():
            mat[0] = f'OpenSees材料-{mat_type}'
            mat[2] = 0
        else:
            mat[0] = self.ui.lineEdit.text()
            mat[2] = 1
        mat[1] = -1  # OpenSees材料的材料种类设为-1
        mat[3] = mat_type
        mat[4] = self.main.mat_N + 1
        mat += mat_para
        if type(self.mod_idx) is int:
            self.main.mat_lib[self.mod_idx] = mat
            self.main.update_conbeBox_mat()
            self.main.ui.listWidget_2.item(self.mod_idx).setText(f'({int(mat[4])}) {mat[0]}')
            print('【Win_OSmat, ok】已修改：', mat)
        else:
            self.main.mat_lib.append(mat)
            self.main.mat_N += 1
            print('【Win_OSmat, ok】已定义：', mat)
            self.main.update_conbeBox_mat()
            self.main.ui.listWidget_2.addItem(f'({int(mat[4])}) {mat[0]}')
        self.accept()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_C and event.modifiers() == Qt.ControlModifier:
            self.copy_selected_cells()
        elif event.key() == Qt.Key_V and event.modifiers() == Qt.ControlModifier:
            self.paste_to_cells()
        else:
            super().keyPressEvent(event)

    def copy_selected_cells(self):
        selected_items = self.ui.tableWidget.selectedItems()
        if not selected_items:
            return
        text = ''
        row_start = selected_items[0].row()
        row_end = selected_items[-1].row()
        col_start = selected_items[0].column()
        col_end = selected_items[-1].column()
        for r in range(row_start, row_end + 1):
            row_data = []
            for c in range(col_start, col_end + 1):
                item = self.ui.tableWidget.item(r, c)
                if item and item.text():
                    row_data.append(item.text())
            text += "\t".join(row_data) + "\n"
        clipboard = QApplication.clipboard()
        clipboard.setText(text)

    def paste_to_cells(self):
        current_row = self.ui.tableWidget.currentRow()
        current_col = self.ui.tableWidget.currentColumn()
        max_row = self.ui.tableWidget.rowCount()
        max_col = self.ui.tableWidget.columnCount()
        clipboard = QApplication.clipboard()
        text = clipboard.text()
        if not text:
            return
        rows = text.strip().split('\n')
        for i, row in enumerate(rows):
            cols = row.split('\t')
            for j, value in enumerate(cols):
                if i + current_row >= max_row or j + current_col >= max_col:
                    continue
                if i + current_row == 1:
                    continue
                item = QTableWidgetItem(value)
                self.ui.tableWidget.setItem(i + current_row, j + current_col, item)
                # self.ui.tableWidget.item(i + current_row, j + current_col).setTextAlignment(0x0004 | 0x0080)

    def showContextMenu(self, pos):
        """定义上下文菜单"""
        context_menu = QMenu(self)
        copy_action = context_menu.addAction("复制")
        paste_action = context_menu.addAction("复制")
        menu_size = context_menu.sizeHint()  # 获取菜单的大小
        global_pos = self.ui.tableWidget.mapToGlobal(pos)  # 获取全局坐标
        adjusted_pos = global_pos + QPoint(0, menu_size.height() // 2)  # 调整位置以使其出现在鼠标指针的右下方
        action = context_menu.exec_(adjusted_pos)
        if action == copy_action:
            self.copy_selected_cells()
        elif action == paste_action:
            self.paste_to_cells()


class Win_setting(QDialog):
    def __init__(self, main: MyWin):
        super().__init__()
        self.ui = Ui_win_solve_setting()
        self.main = main
        self.ui.setupUi(self)
        self.init_ui()
        self.old_values = self.get_value()
    
    def init_ui(self):
        self.ui.lineEdit_5.setPlaceholderText('tol')
        self.ui.lineEdit_6.setPlaceholderText('iter')
        validator = QDoubleValidator()
        self.ui.lineEdit_3.setValidator(validator)
        self.ui.lineEdit_4.setValidator(validator)
        self.ui.lineEdit_5.setValidator(validator)
        self.ui.lineEdit_6.setValidator(validator)
        self.ui.lineEdit_9.setValidator(validator)
        self.ui.lineEdit_10.setValidator(validator)
        self.ui.lineEdit_7.setValidator(validator)
        self.ui.lineEdit.setValidator(validator)
        self.ui.lineEdit_2.setValidator(validator)
        self.ui.comboBox.currentIndexChanged.connect(self.constraint_changed)
        self.ui.comboBox_6.currentIndexChanged.connect(self.integrator_changed)
        self.ui.lineEdit_3.editingFinished.connect(self.value_changed)
        self.ui.lineEdit_4.editingFinished.connect(self.value_changed)
        self.ui.lineEdit_5.editingFinished.connect(self.value_changed)
        self.ui.lineEdit_6.editingFinished.connect(self.value_changed)
        self.ui.lineEdit_7.editingFinished.connect(self.value_changed)
        self.ui.lineEdit_9.editingFinished.connect(self.value_changed)
        self.ui.lineEdit_10.editingFinished.connect(self.value_changed)
        self.ui.lineEdit.editingFinished.connect(self.value_changed)
        self.ui.lineEdit_2.editingFinished.connect(self.value_changed)
        self.ui.pushButton.clicked.connect(self.ok)
        self.ui.pushButton_2.clicked.connect(self.cancel)
        self.ui.pushButton_3.clicked.connect(lambda: self.init_values(self.main.setting_default))
        self.init_values()

    def init_values(self, setting=None):
        if not setting:
            setting = self.main.setting
        self.ui.comboBox.setCurrentIndex(setting[0])
        self.ui.comboBox_2.setCurrentIndex(setting[1])
        self.ui.comboBox_3.setCurrentIndex(setting[2])
        self.ui.comboBox_4.setCurrentIndex(setting[3])
        self.ui.comboBox_5.setCurrentIndex(setting[4])
        self.ui.comboBox_6.setCurrentIndex(setting[5])
        self.ui.lineEdit_3.setText(setting[6])
        self.ui.lineEdit_4.setText(setting[7])
        self.ui.lineEdit_5.setText(setting[8])
        self.ui.lineEdit_6.setText(setting[9])
        self.ui.lineEdit_9.setText(setting[10])
        self.ui.lineEdit_10.setText(setting[11])
        self.ui.lineEdit.setText(setting[12])
        self.ui.lineEdit_2.setText(setting[13])
        self.ui.lineEdit_7.setText(setting[14])

    def get_value(self):
        idx1 = self.ui.comboBox.currentIndex()
        idx2 = self.ui.comboBox_2.currentIndex()
        idx3 = self.ui.comboBox_3.currentIndex()
        idx4 = self.ui.comboBox_4.currentIndex()
        idx5 = self.ui.comboBox_5.currentIndex()
        idx6 = self.ui.comboBox_6.currentIndex()
        cons_val1 = self.ui.lineEdit_3.text()
        cons_val2 = self.ui.lineEdit_4.text()
        text_val1 = self.ui.lineEdit_5.text()
        text_val2 = self.ui.lineEdit_6.text()
        int_val1 = self.ui.lineEdit_9.text()
        int_val2 = self.ui.lineEdit_10.text()
        max_factor = self.ui.lineEdit.text()
        min_factor = self.ui.lineEdit_2.text()
        dt_ratio = self.ui.lineEdit_7.text()
        return [idx1, idx2, idx3, idx4, idx5, idx6, cons_val1, cons_val2, text_val1, text_val2, \
                int_val1, int_val2, max_factor, min_factor, dt_ratio]

    def constraint_changed(self):
        if self.ui.comboBox.currentIndex() == 2:
            self.ui.lineEdit_3.setEnabled(True)
            self.ui.lineEdit_4.setEnabled(True)
            self.ui.lineEdit_3.setPlaceholderText('alphaS')
            self.ui.lineEdit_4.setPlaceholderText('alphaM')
        else:
            self.ui.lineEdit_3.setEnabled(False)
            self.ui.lineEdit_4.setEnabled(False)
            self.ui.lineEdit_3.setPlaceholderText('')
            self.ui.lineEdit_4.setPlaceholderText('')

    def integrator_changed(self):
        if self.ui.comboBox_6.currentIndex() == 1:
            self.ui.lineEdit_9.setEnabled(True)
            self.ui.lineEdit_10.setEnabled(True)
            self.ui.lineEdit_9.setPlaceholderText('gamma')
            self.ui.lineEdit_10.setPlaceholderText('beta') 
        elif self.ui.comboBox_6.currentIndex() == 2:
            self.ui.lineEdit_9.setEnabled(True)
            self.ui.lineEdit_10.setEnabled(False)
            self.ui.lineEdit_9.setPlaceholderText('alpha')
            self.ui.lineEdit_10.setPlaceholderText('')
        elif self.ui.comboBox_6.currentIndex() == 3:
            self.ui.lineEdit_9.setEnabled(True)
            self.ui.lineEdit_10.setEnabled(True)
            self.ui.lineEdit_9.setPlaceholderText('alphaM')
            self.ui.lineEdit_10.setPlaceholderText('alphaF') 
        else:
            self.ui.lineEdit_9.setEnabled(False)
            self.ui.lineEdit_10.setEnabled(False)
            self.ui.lineEdit_9.setPlaceholderText('')
            self.ui.lineEdit_10.setPlaceholderText('')

    def value_changed(self):
        if self.ui.lineEdit_3.text() and float(self.ui.lineEdit_3.text()) < 0:
            self.ui.lineEdit_3.setText('0')
        if self.ui.lineEdit_4.text() and float(self.ui.lineEdit_4.text()) < 0:
            self.ui.lineEdit_4.setText('0')
        if self.ui.lineEdit_5.text() and float(self.ui.lineEdit_5.text()) < 0:
            self.ui.lineEdit_5.setText('0')
        if self.ui.lineEdit_6.text() and float(self.ui.lineEdit_6.text()) < 0:
            self.ui.lineEdit_6.setText('0')
        if self.ui.lineEdit_9.text() and float(self.ui.lineEdit_9.text()) < 0:
            self.ui.lineEdit_9.setText('0')
        if self.ui.lineEdit_10.text() and float(self.ui.lineEdit_10.text()) < 0:
            self.ui.lineEdit_10.setText('0')
        if self.ui.lineEdit.text() and float(self.ui.lineEdit.text()) < 1:
            self.ui.lineEdit.setText('1')
        if self.ui.lineEdit_2.text() and float(self.ui.lineEdit_2.text()) <= 0:
            self.ui.lineEdit_2.setText('1e-15')
        if self.ui.lineEdit_2.text() and float(self.ui.lineEdit_2.text()) > 1:
            self.ui.lineEdit_2.setText('1')
        if self.ui.lineEdit_7.text() and float(self.ui.lineEdit_7.text()) <= 0:
            self.ui.lineEdit_7.setText('1e-6')
        if self.ui.lineEdit_7.text() and float(self.ui.lineEdit_7.text()) > 1:
            self.ui.lineEdit_7.setText('1')

    def check_input(self, *arg):
        """检查文本行是否全部输入数据"""
        ok = True
        for widget in arg:
            if widget.isEnabled():
                if not widget.text():
                    ok = False
                    break
        return ok

    def ok(self):
        if not self.check_input(self.ui.lineEdit_3, self.ui.lineEdit_4, self.ui.lineEdit_5, self.ui.lineEdit_6,\
                                self.ui.lineEdit_9, self.ui.lineEdit_10, self.ui.lineEdit, self.ui.lineEdit_2, self.ui.lineEdit_7):
            QMessageBox.warning(self, '警告', '存在参数未输入！')
            return
        setting = self.get_value()
        self.main.setting = setting
        self.accept()
        print('【Win_setting, ok】设置：\n', setting)

    def cancel(self):
        value = self.old_values
        self.ui.comboBox.setCurrentIndex(value[0])
        self.ui.comboBox_2.setCurrentIndex(value[1])
        self.ui.comboBox_3.setCurrentIndex(value[2])
        self.ui.comboBox_4.setCurrentIndex(value[3])
        self.ui.comboBox_5.setCurrentIndex(value[4])
        self.ui.comboBox_6.setCurrentIndex(value[5])
        self.ui.lineEdit_3.setText(value[6])
        self.ui.lineEdit_4.setText(value[7])
        self.ui.lineEdit_5.setText(value[8])
        self.ui.lineEdit_6.setText(value[9])
        self.ui.lineEdit_9.setText(value[10])
        self.ui.lineEdit_10.setText(value[11])
        self.ui.lineEdit.setText(value[12])
        self.ui.lineEdit_2.setText(value[13])
        self.ui.lineEdit_7.setText(value[14])
        self.accept()


class Win_run(QDialog):
    signal_converge_fail = pyqtSignal()
    signal_finished = pyqtSignal()

    def __init__(self, main: MyWin, script_type):
        super().__init__()
        self.ui = Ui_win_run()
        self.main = main
        self.script_type = script_type
        self.ui.setupUi(self)
        self.init_ui()
        self.start_thread()

    def init_ui(self):
        self.ui.pushButton.clicked.connect(self.click_kill)
        self.setWindowFlag(Qt.WindowCloseButtonHint, False)
        self.ui.label_2.setText(f'正在计算第1条地震动（共{self.main.gm_N}条）')

    def click_kill(self):
        if QMessageBox.question(self, '警告', '是否中断计算？') == QMessageBox.Yes:
            print('【Win_run, click_kill】中断')
            self.kill()

    def kill(self):
        self.thread_run.is_kill = 1

    def start_thread(self):
        self.thread_run = WorkerThread(self.main, self.script_type)
        self.thread_run.signal_finished.connect(self.run_finished)
        self.thread_run.signal_step.connect(self.updata_progressBar)
        self.thread_run.signal_converge.connect(self.is_converge)
        self.thread_run.start()

    def run_finished(self, n):
        """n: 1-正常计算完成，0-计算中断"""
        print('【Win_run, run_finished】计算完成')
        self.accept()
        if n == 1:
            self.main.result_exists = True
            self.signal_finished.emit()

    def updata_progressBar(self, list_):
        n, pct = list_
        self.ui.progressBar.setValue(pct)
        self.ui.label_2.setText(f'正在计算第{n}条地震动（共{self.main.gm_N}条）')
        

    def is_converge(self, list_):
        if list_[0] == 0:
            self.accept()
            QMessageBox.warning(self, '警告', f'地震动{list_[1]}不收敛！')
            self.signal_converge_fail.emit()

    @staticmethod
    def add_free_vibration(th: np.ndarray, fv_time: int | float, dt: float) -> np.ndarray:
        """为时程序列补零"""
        n = int(fv_time / dt) # 补零个数
        th_0 = np.zeros(n)
        th = np.append(th, th_0)
        return th


class WorkerThread(QThread):
    signal_finished = pyqtSignal(int)  # 1: 正常计算完成，0: 计算中断
    signal_step = pyqtSignal(list)
    signal_converge = pyqtSignal(list)  # [n, gm_name], n=1: 收敛，n=0: 不收敛

    def __init__(self, main: MyWin, script_type: str):
        super().__init__()
        self.main = main
        self.script_type = script_type
        self.is_kill = 0

    def run(self):
        gm_N = self.main.gm_N
        for i in range(gm_N):
            print(f'【WorkerThread, run】正在运行...({i+1}/{gm_N})')
            if self.script_type == 'py':
                done = self.solve_py(i)
            else:
                done = self.solve_tcl(i)
            pct = int((i + 1) / gm_N * 100)
            self.signal_step.emit([i + 1, pct])
            if self.is_kill == 1:
                self.signal_finished.emit(0)
                break  # 完成计算
            if done == 0:
                break  # 不收敛
        else:
            self.signal_finished.emit(1)

    def solve_py(self, i):
        print(f'【WorkerThread, solve_py】正在计算第{i+1}条地震动...')
        N = self.main.N
        m = self.main.m
        mat_lib = self.main.mat_lib
        mat_lib = self.check_BW_mat(mat_lib)
        mat_lib = [i[3:] for i in mat_lib]
        story_mat = self.main.story_mat
        th = self.main.gm[i]
        unit = self.main.gm_unit[i]
        if unit == 'g':
            SF = self.main.unit_SF[0]
        elif unit == 'mm/s^2':
            SF = self.main.unit_SF[1]
        elif unit == 'cm/s^2':
            SF = self.main.unit_SF[2]
        elif unit == 'm/s^2':
            SF = self.main.unit_SF[3]
        dt = self.main.gm_dt[i]
        th = Win_run.add_free_vibration(th, self.main.fvtime, dt)
        has_damping = self.main.has_damping
        mode_num = self.main.mode_num
        zeta_mode = self.main.zeta_mode
        zeta = self.main.zeta
        zeta = [float(i) for i in zeta]
        setting = self.main.setting.copy()
        setting[0] = MyWin.setting1[setting[0]]
        setting[1] = MyWin.setting2[setting[1]]
        setting[2] = MyWin.setting3[setting[2]]
        setting[3] = MyWin.setting4[setting[3]]
        setting[4] = MyWin.setting5[setting[4]]
        setting[5] = MyWin.setting6[setting[5]]
        setting[8:] = [eval(i) for i in setting[8:]]
        path = self.main.TEMP_PATH
        gm_name = self.main.gm_name[i]
        done, T, element_tags = core.run_OS_py(
            N, m, mat_lib, story_mat, th, SF, dt, mode_num, has_damping, zeta_mode, zeta, setting, path, gm_name, MyWin.g, MyWin.print_result
        )
        self.signal_converge.emit([done, gm_name])
        return done

    def solve_tcl(self, i):
        print(f'【WorkerThread, solve_tcl】正在计算第{i+1}条地震动')
        N = self.main.N
        m = self.main.m
        mat_lib = self.main.mat_lib
        mat_lib = self.check_BW_mat(mat_lib)
        story_mat = self.main.story_mat
        th = self.main.gm[i]
        unit = self.main.gm_unit[i]
        if unit == 'g':
            SF = self.main.unit_SF[0]
        elif unit == 'mm/s^2':
            SF = self.main.unit_SF[1]
        elif unit == 'cm/s^2':
            SF = self.main.unit_SF[2]
        elif unit == 'm/s^2':
            SF = self.main.unit_SF[3]
        dt = self.main.gm_dt[i]
        th = Win_run.add_free_vibration(th, self.main.fvtime, dt)
        mode_num = self.main.mode_num
        has_damping = self.main.has_damping
        zeta_mode = self.main.zeta_mode
        zeta = self.main.zeta
        setting = self.main.setting
        path = self.main.TEMP_PATH
        path_gm = path + '\\temp_NLMDOF_results\\temp_gm'
        if not os.path.exists(path_gm):
            os.makedirs(path_gm)
        th_path = path_gm + '\\th.txt'
        th_path = th_path.replace('\\', '/')
        np.savetxt(th_path, th)
        gm_name = self.main.gm_name[i]
        NPTS = len(th) - 1
        tcl_script = MyWin.build_tcl_file(
            N, m, mat_lib, story_mat, th_path, SF, dt, mode_num, has_damping, zeta_mode, zeta, setting, path, gm_name, NPTS, MyWin.print_result
        )
        path_tcl = path + '\\temp_NLMDOF_results\\tcl_file'
        if not os.path.exists(path_tcl):
            os.makedirs(path_tcl)
        with open(path_tcl + '\\main.tcl', 'w') as f:
            f.write(tcl_script)
        os.system(f"{self.main.OS_terminal} {path_tcl}\\main.tcl")
        with open(path + '\\temp_NLMDOF_results\\done.txt', 'r') as f:
            if '1' in f.read():
                done = 1
            else:
                done = 0
        self.signal_converge.emit([done, gm_name])
        return done
    
    @staticmethod
    def check_BW_mat(mat_lib):
        """检查是否有boucwen模型，有则替换模型参数格式"""
        # 由Fy格式转为k格式
        new_mat_lib = []
        for _, mat in enumerate(mat_lib):
            if mat[1] == 2:
                # 材料为内置的Wen模型
                new_mat = mat.copy()[:5]
                gamma2, beta2 = 0.5, 0.5
                Fy, uy = float(mat[5]), float(mat[6])
                alpha, n = float(mat[7]), float(mat[8])
                k = Fy / uy
                beta1 = beta2 / uy ** n
                gamma1 = gamma2 / uy ** n
                new_mat.extend([alpha, k, n, gamma1, beta1, 1, 0, 0, 0])
                print('【Win_run, check_BW_mat】更改BoucWen模型格式：', new_mat)
            else:
                new_mat = mat.copy()
            new_mat_lib.append(new_mat)
        return new_mat_lib


class Win_tcl_file(QDialog):
    def __init__(self, text):
        super().__init__()
        self.ui = Ui_win_tcl_file()
        self.ui.setupUi(self)
        self.init_ui(text)

    def init_ui(self, text):
        self.ui.textEdit.setText(text)


class Win_data(QDialog):
    def __init__(self, main: MyWin):
        super().__init__()
        self.ui = Ui_win_data()
        self.main = main
        self.ui.setupUi(self)
        self.init_ui()
        self.init_display_data()

    def init_ui(self):
        self.ui.pushButton.clicked.connect(self.copy_all)
        header = self.ui.tableWidget.horizontalHeader()
        self.ui.tableWidget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.ui.tableWidget.customContextMenuRequested.connect(self.showContextMenu)
        self.ui.tableWidget.setEditTriggers(QTableWidget.NoEditTriggers)
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setStretchLastSection(True)

    def init_display_data(self):
        if not self.main.current_plot_data:
            return
        x, y, title, x_title, y_title = self.main.current_plot_data
        n = len(x)
        self.ui.label.setText(title + '：')
        self.ui.tableWidget.setHorizontalHeaderLabels([x_title, y_title])
        self.ui.tableWidget.setRowCount(n)
        for i, (xi, yi) in enumerate(zip(x, y)):
            xi = round(xi, 7)
            yi = round(yi, 7)
            self.ui.tableWidget.setItem(i, 0, QTableWidgetItem(str(xi)))
            self.ui.tableWidget.setItem(i, 1, QTableWidgetItem(str(yi)))
            self.ui.tableWidget.item(i, 0).setTextAlignment(0x0004 | 0x0080)  # 设置居中
            self.ui.tableWidget.item(i, 1).setTextAlignment(0x0004 | 0x0080)

    def copy_all(self):
        x, y, _, _, _ = self.main.current_plot_data
        text = ''
        for i, (xi, yi) in enumerate(zip(x, y)):
            text += str(round(xi, 7)) + '\t'
            text += str(round(yi, 7)) + '\n'
        clipboard = QApplication.clipboard()
        clipboard.setText(text)
        QMessageBox.information(self, '提示', '已复制。')

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_C and event.modifiers() == Qt.ControlModifier:
            self.copy_selected_cells()
        else:
            super().keyPressEvent(event)

    def copy_selected_cells(self):
        selected_items = self.ui.tableWidget.selectedItems()
        if not selected_items:
            return
        text = ''
        row_start = selected_items[0].row()
        row_end = selected_items[-1].row()
        col_start = selected_items[0].column()
        col_end = selected_items[-1].column()
        for r in range(row_start, row_end + 1):
            row_data = []
            for c in range(col_start, col_end + 1):
                item = self.ui.tableWidget.item(r, c)
                if item and item.text():
                    row_data.append(item.text())
            text += "\t".join(row_data) + "\n"
        clipboard = QApplication.clipboard()
        clipboard.setText(text)

    def showContextMenu(self, pos):
        """定义上下文菜单"""
        context_menu = QMenu(self)
        copy_action = context_menu.addAction("复制")
        menu_size = context_menu.sizeHint()  # 获取菜单的大小
        global_pos = self.ui.tableWidget.mapToGlobal(pos)  # 获取全局坐标
        adjusted_pos = global_pos + QPoint(0, menu_size.height() // 2)  # 调整位置以使其出现在鼠标指针的右下方
        action = context_menu.exec_(adjusted_pos)
        if action == copy_action:
            self.copy_selected_cells()


class Win_export(QDialog):
    def __init__(self, main: MyWin):
        super().__init__()
        self.ui = Ui_win_export()
        self.main = main
        self.ui.setupUi(self)
        self.init_ui()

    def init_ui(self):
        self.ui.pushButton.clicked.connect(self.set_txt)
        self.ui.pushButton_2.clicked.connect(self.set_xlsx)

    def set_xlsx(self):
        self.main.export_type = 'xlsx'
        self.accept()
    
    def set_txt(self):
        self.main.export_type = 'txt'
        self.accept()

class Thread_export_data(QThread):
    signal_finished = pyqtSignal()
    signal_msg = pyqtSignal(list)
    signal_info = pyqtSignal(str)  # 导出数据提示

    def __init__(self, main: MyWin, export_type: str, path: str):
        super().__init__()
        self.main = main
        self.export_type = export_type
        self.path = path

    def run(self):
        if self.export_type == 'xlsx':
            self.export_data_xlsx(self.path)
        else:
            self.export_data_txt(self.path)

    def export_data_txt(self, output_path):
        mode_results = self.main.mode_results
        all_resutls = self.main.all_resutls
        T, mode = mode_results.T, mode_results.mode
        np.savetxt(f'{output_path}/周期.txt', T, fmt='%.7f')
        np.savetxt(f'{output_path}/振型.txt', mode, fmt='%.7f')
        for i, gm_name in enumerate(self.main.gm_name):
            print(f'Thread_export_data, export_data_txt】正在导出计算结果 ({i+1}/{self.main.gm_N})\r', end='')
            results = all_resutls[i]
            self.signal_info.emit(f'正在导出计算结果 ({i+1}/{self.main.gm_N})...')
            t, base_a, base_v, base_u, base_V, aa, av, au, ra, rv, ru, resu, mat = results.all_responses
            np.savetxt(f'{output_path}/{i+1}_{gm_name}_时间序列.txt', t, fmt='%.7f')
            np.savetxt(f'{output_path}/{i+1}_{gm_name}_基底加速度(g).txt', base_a / self.main.g, fmt='%.7f')
            np.savetxt(f'{output_path}/{i+1}_{gm_name}_基底速度(mm_s).txt', base_v, fmt='%.7f')
            np.savetxt(f'{output_path}/{i+1}_{gm_name}_基底位移(mm).txt', base_u, fmt='%.7f')
            np.savetxt(f'{output_path}/{i+1}_{gm_name}_底部剪力(kN).txt', base_V / 1000, fmt='%.7f')
            np.savetxt(f'{output_path}/{i+1}_{gm_name}_绝对加速度(g).txt', aa / self.main.g, fmt='%.7f')
            np.savetxt(f'{output_path}/{i+1}_{gm_name}_绝对速度(mm_s).txt', av, fmt='%.7f')
            np.savetxt(f'{output_path}/{i+1}_{gm_name}_绝对位移(mm).txt', au, fmt='%.7f')
            np.savetxt(f'{output_path}/{i+1}_{gm_name}_相对加速度.txt', ra / self.main.g, fmt='%.7f')
            np.savetxt(f'{output_path}/{i+1}_{gm_name}_相对速度.txt', rv, fmt='%.7f')
            np.savetxt(f'{output_path}/{i+1}_{gm_name}_相对位移.txt', rv, fmt='%.7f')
            interstory_ru = np.diff(ru, axis=1)
            interstory_ru = np.column_stack((ru[:, 0], interstory_ru))
            max_IDR = np.amax(abs(interstory_ru), axis=0)
            np.savetxt(f'{output_path}/{i+1}_{gm_name}_最大层间位移(mm).txt', max_IDR, fmt='%.7f')
            resu = results.resu
            RIDR = np.diff(resu)
            RIDR = np.insert(RIDR, 0, resu[0])
            np.savetxt(f'{output_path}/{i+1}_{gm_name}_相对残余位移(mm).txt', RIDR, fmt='%.7f')
            max_aa = np.max(abs(aa), axis=0)
            np.savetxt(f'{output_path}/{i+1}_{gm_name}_绝对加速度包络(g).txt', max_aa / self.main.g, fmt='%.7f')
            shear = np.zeros((len(t), self.main.N))
            mat_idx: list[list[int]] = []
            n = 0
            for story_idx, sub_list in enumerate(self.main.story_mat):
                mat_idx.append([])
                hys_curve_parallel = np.zeros((len(t), 2))  # 并联材料滞回曲线
                for tag, _ in enumerate(sub_list):
                    mat_idx[story_idx].append(n)
                    col_idx = 2 * n
                    hys_curve = np.zeros((len(t), 2))  # 滞回曲线
                    hys_curve[:, 0] = mat[:, col_idx + 1]
                    hys_curve[:, 1] = mat[:, col_idx] / 1000
                    hys_curve_parallel[:, 0] = mat[:, col_idx + 1]
                    hys_curve_parallel[:, 1] += mat[:, col_idx] / 1000
                    np.savetxt(f'{output_path}/{i+1}_{gm_name}_{story_idx+1}层材料滞回曲线_{self.main.mat_lib[tag-1][0]}.txt', hys_curve, fmt='%.7f')
                    n += 1
                    shear[:, story_idx] += mat[:, col_idx] / 1000
                if len(sub_list) > 1:
                    np.savetxt(f'{output_path}/{i+1}_{gm_name}_{story_idx+1}层并联材料滞回曲线.txt', hys_curve_parallel, fmt='%.7f')
            np.savetxt(f'{output_path}/{i+1}_{gm_name}_楼层剪力(kN).txt', shear, fmt='%.7f')
            shear_enve = np.amax(abs(shear), axis=0)
            np.savetxt(f'{output_path}/{i+1}_{gm_name}_楼层剪力包络(kN).txt', shear_enve, fmt='%.7f')
        with open(f'{output_path}/单位制.txt', 'w') as f:
            f.write('单位：\nN，mm，s\n')
        print(f'Thread_export_data, export_data_txt】已保存计算结果至：{output_path}')
        self.signal_msg.emit(['information', f'已保存计算结果至：\n{output_path}'])
        self.signal_finished.emit()

    def export_data_xlsx(self, excel_path):
        space = ' ' * 20
        print(f'【Thread_export_data, export_data_xlsx】正在导出振型与刚度{space}\r', end='')
        self.signal_info.emit(f'正在导出振型与刚度...')
        mode_results = self.main.mode_results
        all_resutls = self.main.all_resutls
        T, mode = mode_results.T, mode_results.mode
        wb = px.Workbook()
        # 1 振型&周期
        ws: Worksheet = wb.active
        ws.title = '振型与周期'
        ws.cell(1, 1, '振型')
        ws.cell(1, 2, '周期 (s)')
        self.main.write_to_excel(ws, [i for i in range(1, 1 + len(T))], 2, 1)
        self.main.write_to_excel(ws, T, 2, 2)
        ws.cell(1, 4, '楼层')
        ws.merge_cells(start_row=1, start_column=4, end_row=2, end_column=4)
        ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=4 + len(T))
        ws.cell(1, 5, '振型位移 (mm)')
        for i in range(self.main.N):
            ws.cell(3 + i, 4, i + 1)
        for i in range(len(mode)):
            ws.cell(2, 5 + i, f'{i + 1}阶')
        self.main.write_to_excel(ws, np.array(mode).T, 3, 5)
        self.set_ws_center(ws)
        # 2 相对位移
        ws: Worksheet = wb.create_sheet('相对位移')
        for i in range(self.main.gm_N):
            print(f'【Thread_export_data, export_data_xlsx】正在导出相对位移 ({i+1}/{self.main.gm_N}){space}\r', end='')
            self.signal_info.emit(f'正在导出相对位移 ({i+1}/{self.main.gm_N})...')
            results = all_resutls[i]
            t, ru = results.t, results.ru
            ws.merge_cells(start_row=1, start_column=1+(self.main.N+1)*i, end_row=1, end_column=(self.main.N+1)*(i+1))
            ws.cell(2, 1+(self.main.N+1)*i, 't')
            for j in range(self.main.N):
                ws.cell(2, 2+(self.main.N+1)*i+j, f'{j+1}层')
            gm_name = self.main.gm_name[i]
            ws.cell(1, 1+(self.main.N+1)*i, gm_name)
            self.main.write_to_excel(ws, ru, 3, 2+(self.main.N+1)*i)
            self.main.write_to_excel(ws, t, 3, 1+(self.main.N+1)*i)
        self.set_ws_center(ws)
        # 3 相对速度
        ws: Worksheet = wb.create_sheet('相对速度')
        for i in range(self.main.gm_N):
            print(f'【Thread_export_data, export_data_xlsx】正在导出相对速度 ({i+1}/{self.main.gm_N}){space}\r', end='')
            self.signal_info.emit(f'正在导出相对速度 ({i+1}/{self.main.gm_N})...')
            results = all_resutls[i]
            t, rv = results.t, results.rv
            ws.merge_cells(start_row=1, start_column=1+(self.main.N+1)*i, end_row=1, end_column=(self.main.N+1)*(i+1))
            ws.cell(2, 1+(self.main.N+1)*i, 't')
            for j in range(self.main.N):
                ws.cell(2, 2+(self.main.N+1)*i+j, f'{j+1}层')
            gm_name = self.main.gm_name[i]
            ws.cell(1, 1+(self.main.N+1)*i, gm_name)
            self.main.write_to_excel(ws, rv, 3, 2+(self.main.N+1)*i)
            self.main.write_to_excel(ws, t, 3, 1+(self.main.N+1)*i)
        self.set_ws_center(ws)
        # 4 相对加速度
        ws: Worksheet = wb.create_sheet('相对加速度')
        for i in range(self.main.gm_N):
            print(f'【Thread_export_data, export_data_xlsx】正在导出相对加速度 ({i+1}/{self.main.gm_N}){space}\r', end='')
            results = all_resutls[i]
            t, ra = results.t, results.ra
            self.signal_info.emit(f'正在导出相对加速度 ({i+1}/{self.main.gm_N})...')
            ws.merge_cells(start_row=1, start_column=1+(self.main.N+1)*i, end_row=1, end_column=(self.main.N+1)*(i+1))
            ws.cell(2, 1+(self.main.N+1)*i, 't')
            for j in range(self.main.N):
                ws.cell(2, 2+(self.main.N+1)*i+j, f'{j+1}层')
            gm_name = self.main.gm_name[i]
            ws.cell(1, 1+(self.main.N+1)*i, gm_name)
            self.main.write_to_excel(ws, ra, 3, 2+(self.main.N+1)*i)
            self.main.write_to_excel(ws, t, 3, 1+(self.main.N+1)*i)
        self.set_ws_center(ws)
        # 5 最大层间位移
        ws: Worksheet = wb.create_sheet('最大层间位移')
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws.cell(1, 1, '楼层')
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=1+self.main.gm_N)
        ws.cell(1, 2, '最大层间位移')
        for i in range(self.main.N):
            ws.cell(3+i, 1, i+1)
        for i in range(self.main.gm_N):
            print(f'【Thread_export_data, export_data_xlsx】正在导出最大层间位移 ({i+1}/{self.main.gm_N}){space}\r', end='')
            results = all_resutls[i]
            ru = results.ru
            interstory_ru = np.diff(ru, axis=1)
            interstory_ru = np.column_stack((ru[:, 0], interstory_ru))
            max_IDR = np.amax(abs(interstory_ru), axis=0)
            self.signal_info.emit(f'正在导出最大层间位移 ({i+1}/{self.main.gm_N})...')
            gm_name = self.main.gm_name[i]
            ws.cell(2, 2+i, gm_name)
            self.main.write_to_excel(ws, max_IDR, 3, 2+i)
        self.set_ws_center(ws)
        # 6 最大层间残余位移
        ws: Worksheet = wb.create_sheet('最大层间残余位移')
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws.cell(1, 1, '楼层')
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=1+self.main.gm_N)
        ws.cell(1, 2, '最大层间残余位移')
        for i in range(self.main.N):
            ws.cell(3+i, 1, i+1)
        for i in range(self.main.gm_N):
            print(f'【Thread_export_data, export_data_xlsx】正在导出最大层间残余位移 ({i+1}/{self.main.gm_N}){space}\r', end='')
            self.signal_info.emit(f'正在导出最大层间残余位移 ({i+1}/{self.main.gm_N})...')
            results = all_resutls[i]
            resu = results.resu
            RIDR = np.diff(resu)
            RIDR = np.insert(RIDR, 0, resu[0])
            RIDR = np.insert(RIDR, 0, 0)
            gm_name = self.main.gm_name[i]
            ws.cell(2, 2+i, gm_name)
            self.main.write_to_excel(ws, resu, 3, 2+i)
        self.set_ws_center(ws)
        # 7 绝对位移
        ws: Worksheet = wb.create_sheet('绝对位移')
        for i in range(self.main.gm_N):
            print(f'【Thread_export_data, export_data_xlsx】正在导出绝对位移 ({i+1}/{self.main.gm_N}){space}\r', end='')
            self.signal_info.emit(f'正在导出绝对位移 ({i+1}/{self.main.gm_N})...')
            results = all_resutls[i]
            t, au = results.t, results.au
            ws.merge_cells(start_row=1, start_column=1+(self.main.N+1)*i, end_row=1, end_column=(self.main.N+1)*(i+1))
            ws.cell(2, 1+(self.main.N+1)*i, 't')
            for j in range(self.main.N):
                ws.cell(2, 2+(self.main.N+1)*i+j, f'{j+1}层')
            gm_name = self.main.gm_name[i]
            ws.cell(1, 1+(self.main.N+1)*i, gm_name)
            self.main.write_to_excel(ws, t, 3, 1+(self.main.N+1)*i)
            self.main.write_to_excel(ws, au, 3, 2+(self.main.N+1)*i)
        self.set_ws_center(ws)
        # 8 绝对速度
        ws: Worksheet = wb.create_sheet('绝对速度')
        for i in range(self.main.gm_N):
            print(f'【Thread_export_data, export_data_xlsx】正在导出绝对速度 ({i+1}/{self.main.gm_N}){space}\r', end='')
            self.signal_info.emit(f'正在导出绝对速度 ({i+1}/{self.main.gm_N})...')
            results = all_resutls[i]
            t, av = results.t, results.av
            ws.merge_cells(start_row=1, start_column=1+(self.main.N+1)*i, end_row=1, end_column=(self.main.N+1)*(i+1))
            ws.cell(2, 1+(self.main.N+1)*i, 't')
            for j in range(self.main.N):
                ws.cell(2, 2+(self.main.N+1)*i+j, f'{j+1}层')
            gm_name = self.main.gm_name[i]
            ws.cell(1, 1+(self.main.N+1)*i, gm_name)
            self.main.write_to_excel(ws, t, 3, 1+(self.main.N+1)*i)
            self.main.write_to_excel(ws, av, 3, 2+(self.main.N+1)*i)
        self.set_ws_center(ws)
        # 9 绝对加速度
        ws: Worksheet = wb.create_sheet('绝对加速度')
        for i in range(self.main.gm_N):
            print(f'【Thread_export_data, export_data_xlsx】正在导出绝对加速度 ({i+1}/{self.main.gm_N}){space}\r', end='')
            self.signal_info.emit(f'正在导出绝对加速度 ({i+1}/{self.main.gm_N})...')
            results = all_resutls[i]
            t, aa = results.t, results.aa / self.main.g
            ws.merge_cells(start_row=1, start_column=1+(self.main.N+1)*i, end_row=1, end_column=(self.main.N+1)*(i+1))
            ws.cell(2, 1+(self.main.N+1)*i, 't')
            for j in range(self.main.N):
                ws.cell(2, 2+(self.main.N+1)*i+j, f'{j+1}层')
            gm_name = self.main.gm_name[i]
            ws.cell(1, 1+(self.main.N+1)*i, gm_name)
            self.main.write_to_excel(ws, t, 3, 1+(self.main.N+1)*i)
            self.main.write_to_excel(ws, aa, 3, 2+(self.main.N+1)*i)
        self.set_ws_center(ws)
        # 10 绝对加速度包络
        ws: Worksheet = wb.create_sheet('绝对加速度包络')
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws.cell(1, 1, '楼层')
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=1+self.main.gm_N)
        ws.cell(1, 2, '绝对加速度包络')
        for i in range(self.main.N):
            ws.cell(3+i, 1, i+1)
        for i in range(self.main.gm_N):
            print(f'【Thread_export_data, export_data_xlsx】正在导出绝对加速度包络 ({i+1}/{self.main.gm_N}){space}\r', end='')
            self.signal_info.emit(f'正在导出绝对加速度包络 ({i+1}/{self.main.gm_N})...')
            results = all_resutls[i]
            t, aa = results.t, results.aa / self.main.g
            max_aa = np.max(abs(aa), axis=0)
            gm_name = self.main.gm_name[i]
            ws.cell(2, 2+i, gm_name)
            self.main.write_to_excel(ws, max_aa, 3, 2+i)
        self.set_ws_center(ws)
        # 11 底部剪力
        ws: Worksheet = wb.create_sheet('底部剪力')
        for i in range(self.main.gm_N):
            print(f'【Thread_export_data, export_data_xlsx】正在导出底部剪力 ({i+1}/{self.main.gm_N}){space}\r', end='')
            self.signal_info.emit(f'正在导出底部剪力 ({i+1}/{self.main.gm_N})...')
            results = all_resutls[i]
            t, base_V = results.t, results.base_V / 1000
            ws.cell(2, 2*i+1, 't')
            ws.cell(2, 2*i+2, 'Vb')
            gm_name = self.main.gm_name[i]
            ws.merge_cells(start_row=1, start_column=2*i+1, end_row=1, end_column=2*i+2)
            ws.cell(1, 2*i+1, gm_name)
            self.main.write_to_excel(ws, t, 3, 2*i+1)
            self.main.write_to_excel(ws, base_V, 3, 2*i+2)
        self.set_ws_center(ws)
        # 12-14 楼层剪力、楼层剪力包络、材料滞回曲线
        ws: Worksheet = wb.create_sheet('材料滞回曲线')
        ws1 = wb.create_sheet('楼层剪力')
        ws2 = wb.create_sheet('楼层剪力包络')
        ws2.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws2.cell(1, 1, '楼层')
        ws2.merge_cells(start_row=1, start_column=2, end_row=1, end_column=1+self.main.gm_N)
        ws2.cell(1, 2, '楼层剪力包络')
        self.main.write_to_excel(ws2, [i+1 for i in range(self.main.N)], 3, 1)
        current_col = 1  # 当前写入的列序号
        current_col_gm = current_col  # 当前地震动的列序号
        for i in range(self.main.gm_N):
            print(f'【Thread_export_data, export_data_xlsx】正在导出楼层剪力、楼层剪力包络、材料滞回曲线 ({i+1}/{self.main.gm_N}){space}\r', end='')
            self.signal_info.emit(f'正在导出楼层剪力、楼层剪力包络、材料滞回曲线 ({i+1}/{self.main.gm_N})...')
            results = all_resutls[i]
            t = results.t
            base_V = results.base_V / 1000
            mat = results.mat
            mat_idx: list[list[int]] = []  # 各层材料的序号
            n = 0
            gm_name = self.main.gm_name[i]
            shear = np.zeros((len(t), self.main.N))
            ws1.merge_cells(start_row=1, start_column=(self.main.N+1)*i+1, end_row=1, end_column=(self.main.N+1)*(i+1))
            ws1.cell(1, (self.main.N+1)*i+1, f'{gm_name}')
            ws1.cell(2, (self.main.N+1)*i+1, 't')
            self.main.write_to_excel(ws1, t, 3, (self.main.N+1)*i+1)
            ws2.cell(2, 2+i, gm_name)
            for story_idx, sub_list in enumerate(self.main.story_mat):
                current_col_u = current_col  # 位移数据所在列序号
                current_col += 1
                mat_idx.append([])
                hys_curve_parallel = np.zeros((len(t), 2))  # 并联材料滞回曲线
                current_col_story = current_col
                ws1.cell(2, (self.main.N+1)*i+2+story_idx, f'{story_idx+1}层')
                for tag, _ in enumerate(sub_list):
                    mat_idx[story_idx].append(n)
                    col_idx = 2 * n
                    hys_curve = np.zeros((len(t), 2))  # 滞回曲线
                    hys_curve[:, 0] = mat[:, col_idx + 1]
                    hys_curve[:, 1] = mat[:, col_idx] / 1000
                    hys_curve_parallel[:, 0] = mat[:, col_idx + 1]
                    hys_curve_parallel[:, 1] += mat[:, col_idx] / 1000
                    self.main.write_to_excel(ws, hys_curve[:, 1], 4, current_col)
                    self.main.write_to_excel(ws, hys_curve[:, 0], 4, current_col_u)
                    ws.cell(3, current_col, self.main.mat_lib[tag-1][0])
                    current_col += 1
                    n += 1
                    shear[:, story_idx] += mat[:, col_idx] / 1000
                if len(sub_list) > 1:
                    self.main.write_to_excel(ws, hys_curve_parallel[:, 1], 4, current_col)
                    ws.cell(3, current_col, '并联材料')
                    current_col += 1
                ws.merge_cells(start_row=2, start_column=current_col_story-1, end_row=2, end_column=current_col-1)
                ws.cell(2, current_col_story-1, f'{story_idx+1}层')
                ws.cell(3, current_col_story-1, 'u')
            self.main.write_to_excel(ws1, shear, 3, (self.main.N+1)*i+2)
            ws.merge_cells(start_row=1, start_column=current_col_gm, end_row=1, end_column=current_col-1)
            ws.cell(1, current_col_gm, f'{gm_name}')
            shear_enve = np.amax(abs(shear), axis=0)
            self.main.write_to_excel(ws2, shear_enve, 3, 2+i)
            current_col_gm = current_col
        self.set_ws_center(ws)
        # 15 单位制
        ws: Worksheet = wb.create_sheet('单位制')
        ws.cell(1, 1, '时间')
        ws.cell(1, 2, 's')
        ws.cell(2, 1, '位移')
        ws.cell(2, 2, 'mm')
        ws.cell(3, 1, '速度')
        ws.cell(3, 2, 'mm/s')
        ws.cell(4, 1, '加速度')
        ws.cell(4, 2, 'g')
        ws.cell(5, 1, '剪力')
        ws.cell(5, 2, 'kN')
        ws.cell(6, 1, f'g = {self.main.g}mm/s^2')
        ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=2)
        self.set_ws_center(ws)
        self.set_ws_center(ws1)
        self.set_ws_center(ws2)
        # 保存文件
        try:
            print(f'【Thread_export_data, export_data_xlsx】正在保存文件...')
            self.signal_info.emit('正在保存文件...')
            wb.save(excel_path)
            print(f'【Thread_export_data, export_data_xlsx】已保存计算结果至：{excel_path}')
            self.signal_msg.emit(['information', f'已保存计算结果至：\n{excel_path}'])
        except:
            self.signal_msg.emit(['critical', '无法保存文件，检查文件是否处于打开状态！'])
        self.signal_finished.emit()

    def set_ws_center(self, ws):
        if self.main.gm_N * self.main.N > 8:
            return  # 如果地震动和层数过多，则不进行居中以节省时间
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = px.styles.Alignment(horizontal='center', vertical='center')


class Win_about(QDialog):
    def __init__(self):
        super().__init__()
        self.ui = Ui_win_about()
        self.ui.setupUi(self)
        self.init_ui()

    def init_ui(self):
        text = self.ui.label_3.text()
        text = text.replace('date', DATE)
        text = text.replace('Version', VERSION)
        self.ui.label_3.setText(text)

